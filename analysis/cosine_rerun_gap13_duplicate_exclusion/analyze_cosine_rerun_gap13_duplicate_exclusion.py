"""Rebuild cross-group cosine figures with step K=13.

Reconstructs the step fixed-K clustering from baseline exports,
removes only duplicate step components within duplicated trial-cluster pairs,
and regenerates cross-group cosine similarity artifacts in `analysis/`.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import polars as pl
import yaml
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import matplotlib

try:
    matplotlib.use("Agg", force=True)
except Exception:
    pass
from matplotlib import font_manager

from src.synergy_stats.clustering import _fit_best_kmeans_result
from src.synergy_stats.cross_group_similarity import (
    annotate_pairwise_assignment,
    build_cluster_decision,
    build_cluster_w_matrix,
    compute_pairwise_cosine,
    solve_assignment,
)
from src.synergy_stats.figures import (
    save_cross_group_decision_summary,
    save_cross_group_heatmap,
    save_cross_group_matched_h,
    save_cross_group_matched_w,
)


STEP_GROUP_ID = "global_step"
NONSTEP_GROUP_ID = "global_nonstep"

W_BASE_COLUMNS = {
    "group_id",
    "subject",
    "velocity",
    "trial_num",
    "trial_id",
    "component_index",
    "muscle",
    "W_value",
}
H_BASE_COLUMNS = {
    "group_id",
    "subject",
    "velocity",
    "trial_num",
    "trial_id",
    "component_index",
    "frame_idx",
    "h_value",
}
KOREAN_FONT_CANDIDATES = (
    "NanumGothic",
    "NanumBarunGothic",
    "Malgun Gothic",
    "AppleGothic",
    "Noto Sans CJK KR",
    "Noto Sans KR",
)


@dataclass
class TrialFeature:
    """One baseline-exported trial reconstructed from minimal W/H tables."""

    subject: str
    velocity: float
    trial_num: int
    trial_id: str
    component_ids: list[int]
    w_muscle: np.ndarray
    h_time: np.ndarray
    meta: dict[str, Any]


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments for dry-run and full rerun modes."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--baseline-run",
        type=Path,
        default=REPO_ROOT / "outputs" / "runs" / "default_run",
        help="Baseline pipeline output directory used as the offline input bundle.",
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=REPO_ROOT / "configs" / "synergy_stats_config.yaml",
        help="Synergy stats config used for muscle order, threshold, and figure settings.",
    )
    parser.add_argument(
        "--outdir",
        type=Path,
        default=SCRIPT_DIR / "artifacts" / "gap13_duplicate_component_exclusion_rerun",
        help="Output directory for regenerated artifacts.",
    )
    parser.add_argument("--step-k", type=int, default=13, help="Fixed step clustering K to rerun.")
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite the output directory if it already exists.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Only validate inputs and reconstruction consistency; skip exports.",
    )
    parser.add_argument(
        "--objective-atol",
        type=float,
        default=0.05,
        help="Absolute tolerance allowed when comparing rerun observed objective against baseline metadata.",
    )
    return parser.parse_args()


def _configure_fonts() -> None:
    """Prefer a Korean-capable font so saved figures keep readable labels."""
    available_fonts = {font.name for font in font_manager.fontManager.ttflist}
    selected = next((name for name in KOREAN_FONT_CANDIDATES if name in available_fonts), None)
    if selected is not None:
        matplotlib.rcParams["font.family"] = [selected]
    matplotlib.rcParams["axes.unicode_minus"] = False


def _ensure_outdir(path: Path, *, overwrite: bool) -> None:
    """Create a clean output directory for the analysis run."""
    if path.exists():
        if not overwrite:
            raise FileExistsError(f"Output directory exists: {path} (use --overwrite)")
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def _load_yaml(path: Path) -> dict[str, Any]:
    """Load the synergy stats YAML config as a plain dict."""
    with path.open("r", encoding="utf-8-sig") as handle:
        return yaml.safe_load(handle)


def _read_csv(path: Path) -> pl.DataFrame:
    """Load a CSV with BOM-friendly settings via Polars first."""
    return pl.read_csv(path, encoding="utf8-lossy", infer_schema_length=10000)


def _to_float(value: Any) -> float:
    """Normalize numeric-like values used as trial keys."""
    return float(value)


def _to_int(value: Any) -> int:
    """Normalize integer-like values used as trial keys."""
    return int(float(value))


def _trial_key(subject: Any, velocity: Any, trial_num: Any) -> tuple[str, float, int]:
    """Build a canonical trial key used across W/H tables."""
    return (str(subject).strip(), _to_float(velocity), _to_int(trial_num))


def _load_trial_features(
    w_long: pl.DataFrame,
    h_long: pl.DataFrame,
    group_id: str,
    muscle_order: list[str],
) -> list[TrialFeature]:
    """Reconstruct per-trial W/H matrices from baseline minimal export tables."""
    w_group = w_long.filter(pl.col("group_id") == group_id).to_pandas()
    h_group = h_long.filter(pl.col("group_id") == group_id).to_pandas()
    meta_columns = [column for column in w_group.columns if column not in W_BASE_COLUMNS]

    trial_order = (
        w_group.loc[:, ["subject", "velocity", "trial_num", "trial_id"]]
        .drop_duplicates(keep="first")
        .reset_index(drop=True)
    )

    features: list[TrialFeature] = []
    for row in trial_order.itertuples(index=False):
        mask_w = (
            (w_group["subject"] == row.subject)
            & (w_group["velocity"].astype(float) == float(row.velocity))
            & (w_group["trial_num"].astype(float) == float(row.trial_num))
        )
        trial_w = w_group.loc[mask_w].copy()
        component_ids = sorted(int(value) for value in trial_w["component_index"].drop_duplicates().tolist())
        w_pivot = trial_w.pivot(index="muscle", columns="component_index", values="W_value")
        w_pivot = w_pivot.reindex(index=muscle_order, columns=component_ids)
        if w_pivot.isnull().values.any():
            raise ValueError(f"Null W values found while reconstructing {row.trial_id}.")

        mask_h = (
            (h_group["subject"] == row.subject)
            & (h_group["velocity"].astype(float) == float(row.velocity))
            & (h_group["trial_num"].astype(float) == float(row.trial_num))
        )
        trial_h = h_group.loc[mask_h].copy()
        h_pivot = trial_h.pivot(index="frame_idx", columns="component_index", values="h_value")
        h_pivot = h_pivot.sort_index().reindex(columns=component_ids)
        if h_pivot.isnull().values.any():
            raise ValueError(f"Null H values found while reconstructing {row.trial_id}.")

        meta_row = trial_w.iloc[0]
        meta = {column: meta_row[column] for column in meta_columns}
        features.append(
            TrialFeature(
                subject=str(row.subject).strip(),
                velocity=_to_float(row.velocity),
                trial_num=_to_int(row.trial_num),
                trial_id=str(row.trial_id),
                component_ids=component_ids,
                w_muscle=w_pivot.to_numpy(dtype=np.float32),
                h_time=h_pivot.to_numpy(dtype=np.float32),
                meta=meta,
            )
        )
    return features


def _stack_weight_vectors(features: list[TrialFeature]) -> tuple[np.ndarray, list[dict[str, Any]]]:
    """Match the pipeline stacking order: trial order first, component order second."""
    vectors: list[np.ndarray] = []
    sample_map: list[dict[str, Any]] = []
    for feature in features:
        trial_key = _trial_key(feature.subject, feature.velocity, feature.trial_num)
        for local_index, component_id in enumerate(feature.component_ids):
            vectors.append(feature.w_muscle[:, local_index].astype(np.float32))
            sample_map.append(
                {
                    "group_id": STEP_GROUP_ID,
                    "subject": feature.subject,
                    "velocity": feature.velocity,
                    "trial_num": feature.trial_num,
                    "trial_id": feature.trial_id,
                    "trial_key": trial_key,
                    "component_index": int(component_id),
                    "local_component_index": int(local_index),
                    "meta": feature.meta,
                }
            )
    if not vectors:
        raise ValueError("No step feature vectors were reconstructed from baseline exports.")
    return np.stack(vectors, axis=0), sample_map


def _stack_weight_vectors_from_baseline_order(
    w_long: pl.DataFrame,
    baseline_labels: pl.DataFrame,
    muscle_order: list[str],
) -> tuple[np.ndarray, list[dict[str, Any]]]:
    """Rebuild the exact original sample_map order from all_cluster_labels.csv."""
    w_group = w_long.filter(pl.col("group_id") == STEP_GROUP_ID).to_pandas()
    baseline_order = baseline_labels.filter(pl.col("group_id") == STEP_GROUP_ID).to_pandas()
    meta_columns = [column for column in baseline_order.columns if column not in {"group_id", "cluster_id"}]

    vector_lookup: dict[tuple[str, int], np.ndarray] = {}
    for (trial_id, component_index), trial_component in w_group.groupby(["trial_id", "component_index"], sort=False):
        ordered = (
            trial_component.loc[:, ["muscle", "W_value"]]
            .drop_duplicates(subset=["muscle"], keep="first")
            .set_index("muscle")
            .reindex(muscle_order)
        )
        if ordered["W_value"].isnull().any():
            raise ValueError(f"Null W values found while building vector lookup for {trial_id}, component {component_index}.")
        vector_lookup[(str(trial_id), int(component_index))] = ordered["W_value"].to_numpy(dtype=np.float32)

    vectors: list[np.ndarray] = []
    sample_map: list[dict[str, Any]] = []
    for row in baseline_order.itertuples(index=False):
        lookup_key = (str(row.trial_id), int(row.component_index))
        if lookup_key not in vector_lookup:
            raise KeyError(f"Missing component vector for baseline order row: {lookup_key}")
        vectors.append(vector_lookup[lookup_key])
        row_meta = {column: getattr(row, column) for column in meta_columns if column not in W_BASE_COLUMNS}
        sample_map.append(
            {
                "group_id": STEP_GROUP_ID,
                "subject": str(row.subject).strip(),
                "velocity": _to_float(row.velocity),
                "trial_num": _to_int(row.trial_num),
                "trial_id": str(row.trial_id),
                "trial_key": _trial_key(row.subject, row.velocity, row.trial_num),
                "component_index": int(row.component_index),
                "local_component_index": int(row.component_index),
                "meta": row_meta,
            }
        )
    if not vectors:
        raise ValueError("No step component vectors were reconstructed from baseline label order.")
    return np.stack(vectors, axis=0), sample_map


def _parse_step_metadata(
    step_metadata: pl.DataFrame,
    baseline_run: Path,
    step_k: int,
) -> tuple[dict[str, Any], dict[str, float], pd.DataFrame]:
    """Read stored K-audit metadata for consistency checks against the rerun."""
    step_row = step_metadata.filter(pl.col("group_id") == STEP_GROUP_ID).to_pandas().iloc[0]
    observed_by_k = {int(key): float(value) for key, value in json.loads(step_row["observed_objective_by_k_json"]).items()}
    duplicate_by_k = {int(key): int(value) for key, value in json.loads(step_row["duplicate_trial_count_by_k_json"]).items()}
    expected_duplicate_evidence = _load_expected_duplicate_evidence(baseline_run, step_k)
    metadata_summary = {
        "k_gap_raw": int(step_row["k_gap_raw"]),
        "k_selected": int(step_row["k_selected"]),
        "k_min_unique": float(step_row["k_min_unique"]),
        "selection_status": str(step_row["selection_status"]),
        "expected_duplicate_trial_count_at_step_k": int(duplicate_by_k.get(step_k, -1)),
        "expected_observed_objective_at_step_k": float(observed_by_k[step_k]),
    }
    return metadata_summary, observed_by_k, expected_duplicate_evidence


def _load_expected_duplicate_evidence(baseline_run: Path, step_k: int) -> pd.DataFrame:
    """Extract exact duplicate cluster/component evidence from the audit workbook."""
    workbook = load_workbook(baseline_run / "clustering_audit.xlsx", data_only=True)
    sheet = workbook["duplicates"]
    detail_headers = [
        "group_id",
        "k",
        "subject",
        "velocity",
        "trial_num",
        "trial_id",
        "cluster_id",
        "component_indexes_json",
        "component_count",
        "is_gap_raw_k",
        "is_selected_k",
    ]
    header_row = None
    for row_index in range(1, sheet.max_row + 1):
        row_values = [sheet.cell(row_index, column_index).value for column_index in range(1, len(detail_headers) + 1)]
        if row_values == detail_headers:
            header_row = row_index
            break
    if header_row is None:
        raise ValueError("Could not find duplicate cluster detail header in clustering_audit.xlsx.")

    rows: list[dict[str, Any]] = []
    row_index = header_row + 1
    while row_index <= sheet.max_row:
        values = [sheet.cell(row_index, column_index).value for column_index in range(1, len(detail_headers) + 1)]
        if values[0] is None:
            break
        if values == detail_headers:
            row_index += 1
            continue
        rows.append(dict(zip(detail_headers, values, strict=True)))
        row_index += 1

    detail_frame = pd.DataFrame(rows)
    detail_frame = detail_frame[(detail_frame["group_id"] == STEP_GROUP_ID) & (detail_frame["k"] == step_k)].copy()
    if detail_frame.empty:
        raise ValueError(f"No duplicate evidence found for {STEP_GROUP_ID} at K={step_k}.")
    detail_frame["component_indexes_json"] = detail_frame["component_indexes_json"].apply(
        lambda value: json.dumps(sorted(json.loads(value)), ensure_ascii=False)
    )
    return detail_frame.loc[:, ["trial_id", "cluster_id", "component_indexes_json"]].sort_values(
        ["trial_id", "cluster_id"]
    ).reset_index(drop=True)


def _compute_centroids(data: np.ndarray, labels: np.ndarray, n_clusters: int) -> np.ndarray:
    """Compute cluster centroids from the fixed-K solution for component filtering."""
    centroids = np.zeros((n_clusters, data.shape[1]), dtype=np.float64)
    for cluster_id in range(n_clusters):
        members = data[labels == cluster_id]
        if len(members) == 0:
            continue
        centroids[cluster_id] = members.mean(axis=0, dtype=np.float64)
    return centroids


def _trial_cluster_groups(labels: np.ndarray, sample_map: list[dict[str, Any]]) -> dict[tuple[str, int], list[int]]:
    """Group sample indices by trial and assigned cluster."""
    groups: dict[tuple[str, int], list[int]] = {}
    for sample_index, (sample, cluster_id) in enumerate(zip(sample_map, labels.tolist(), strict=True)):
        groups.setdefault((sample["trial_id"], int(cluster_id)), []).append(sample_index)
    return groups


def _duplicate_trial_ids(labels: np.ndarray, sample_map: list[dict[str, Any]]) -> list[str]:
    """List trial IDs that still contain within-trial duplicate cluster assignments."""
    duplicate_trials = {
        trial_id
        for (trial_id, _cluster_id), member_indices in _trial_cluster_groups(labels, sample_map).items()
        if len(member_indices) > 1
    }
    return sorted(duplicate_trials)


def _select_min_duplicate_candidate(
    data: np.ndarray,
    sample_map: list[dict[str, Any]],
    n_clusters: int,
    clustering_cfg: dict[str, Any],
    observed_result: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Mirror the uniqueness-search path and keep the least-duplicate fixed-K candidate."""
    candidate_restarts = int(clustering_cfg.get("uniqueness_candidate_restarts", clustering_cfg.get("repeats", 25)))
    base_seed = int(clustering_cfg.get("random_state", 42)) + (int(n_clusters) * 100000)
    best: dict[str, Any] | None = None
    if observed_result is not None:
        observed_labels = np.asarray(observed_result["labels"], dtype=np.int32)
        best = {
            "labels": observed_labels,
            "objective": float(observed_result["objective"]),
            "seed": int(base_seed),
            "restart_index": -1,
            "duplicate_trials": _duplicate_trial_ids(observed_labels, sample_map),
            "duplicate_trial_count": int(len(_duplicate_trial_ids(observed_labels, sample_map))),
        }
    for restart_index in range(candidate_restarts):
        candidate_seed = base_seed + restart_index
        candidate = _fit_best_kmeans_result(
            data,
            int(n_clusters),
            1,
            candidate_seed,
            clustering_cfg,
        )
        labels = np.asarray(candidate["labels"], dtype=np.int32)
        duplicate_trials = _duplicate_trial_ids(labels, sample_map)
        payload = {
            "labels": labels,
            "objective": float(candidate["objective"]),
            "seed": int(candidate_seed),
            "restart_index": int(restart_index),
            "duplicate_trials": duplicate_trials,
            "duplicate_trial_count": int(len(duplicate_trials)),
        }
        if best is None:
            best = payload
            continue
        if payload["duplicate_trial_count"] < best["duplicate_trial_count"]:
            best = payload
            continue
        if (
            payload["duplicate_trial_count"] == best["duplicate_trial_count"]
            and payload["objective"] < best["objective"]
        ):
            best = payload
    if best is None:
        raise ValueError(f"Could not find a fixed-K candidate at K={n_clusters}.")
    return best


def _normalize_duplicate_evidence(frame: pd.DataFrame, component_column: str) -> pd.DataFrame:
    """Normalize duplicate evidence tables for exact equality checks."""
    normalized = frame.copy()
    normalized[component_column] = normalized[component_column].apply(
        lambda value: json.dumps(sorted(json.loads(value)), ensure_ascii=False)
    )
    normalized["cluster_id"] = normalized["cluster_id"].astype(int)
    normalized["trial_id"] = normalized["trial_id"].astype(str)
    return normalized.loc[:, ["trial_id", "cluster_id", component_column]].sort_values(
        ["trial_id", "cluster_id"]
    ).reset_index(drop=True)


def _cosine_similarity(left: np.ndarray, right: np.ndarray) -> float:
    """Return cosine similarity with safe handling for degenerate vectors."""
    left_norm = float(np.linalg.norm(left))
    right_norm = float(np.linalg.norm(right))
    if left_norm <= 0.0 or right_norm <= 0.0:
        return 0.0
    return float(np.clip(np.dot(left, right) / (left_norm * right_norm), -1.0, 1.0))


def _duplicate_component_decisions(
    data: np.ndarray,
    labels: np.ndarray,
    sample_map: list[dict[str, Any]],
    n_clusters: int,
) -> tuple[pd.DataFrame, pd.DataFrame, set[int]]:
    """Keep one component per duplicated trial-cluster pair and exclude the rest."""
    centroids = _compute_centroids(data, labels, n_clusters)
    rows: list[dict[str, Any]] = []
    summary_rows: list[dict[str, Any]] = []
    excluded_indices: set[int] = set()

    trial_cluster_groups = _trial_cluster_groups(labels, sample_map)

    keep_index_by_group: dict[tuple[str, int], int] = {}
    for (trial_id, cluster_id), member_indices in trial_cluster_groups.items():
        if len(member_indices) <= 1:
            keep_index_by_group[(trial_id, cluster_id)] = member_indices[0]
            continue
        scored = []
        for sample_index in member_indices:
            sample = sample_map[sample_index]
            cosine_to_centroid = _cosine_similarity(data[sample_index], centroids[cluster_id])
            scored.append((cosine_to_centroid, -int(sample["component_index"]), sample_index))
        best = max(scored)
        keep_index = int(best[2])
        keep_index_by_group[(trial_id, cluster_id)] = keep_index
        excluded_here = sorted(
            int(sample_map[idx]["component_index"])
            for idx in member_indices
            if idx != keep_index
        )
        summary_rows.append(
            {
                "trial_id": trial_id,
                "subject": sample_map[member_indices[0]]["subject"],
                "velocity": sample_map[member_indices[0]]["velocity"],
                "trial_num": sample_map[member_indices[0]]["trial_num"],
                "cluster_id": cluster_id,
                "component_indexes_in_duplicate_set": json.dumps(
                    sorted(int(sample_map[idx]["component_index"]) for idx in member_indices),
                    ensure_ascii=False,
                ),
                "kept_component_index": int(sample_map[keep_index]["component_index"]),
                "excluded_component_indexes": json.dumps(excluded_here, ensure_ascii=False),
                "duplicate_component_count": len(member_indices),
            }
        )
        for sample_index in member_indices:
            if sample_index != keep_index:
                excluded_indices.add(sample_index)

    for sample_index, (sample, cluster_id) in enumerate(zip(sample_map, labels.tolist(), strict=True)):
        group_key = (sample["trial_id"], int(cluster_id))
        cosine_to_centroid = _cosine_similarity(data[sample_index], centroids[int(cluster_id)])
        duplicate_count = len(trial_cluster_groups[group_key])
        keep_index = keep_index_by_group[group_key]
        is_duplicate_group = duplicate_count > 1
        exclude = sample_index in excluded_indices
        rows.append(
            {
                "group_id": sample["group_id"],
                "subject": sample["subject"],
                "velocity": sample["velocity"],
                "trial_num": sample["trial_num"],
                "trial_id": sample["trial_id"],
                "component_index": int(sample["component_index"]),
                "cluster_id_k13": int(cluster_id),
                "cosine_to_cluster_centroid": cosine_to_centroid,
                "is_duplicate_group": is_duplicate_group,
                "duplicate_component_count_in_trial_cluster": duplicate_count,
                "kept_in_duplicate_group": bool(sample_index == keep_index),
                "excluded_from_representative": bool(exclude),
                "exclusion_reason": (
                    "duplicate_component_removed"
                    if exclude
                    else ("kept_best_centroid_match" if is_duplicate_group else "unique_assignment")
                ),
                **sample["meta"],
            }
        )
    component_table = pd.DataFrame(rows).sort_values(["trial_id", "component_index"]).reset_index(drop=True)
    duplicate_summary = pd.DataFrame(summary_rows).sort_values(["trial_id", "cluster_id"]).reset_index(drop=True)
    return component_table, duplicate_summary, excluded_indices


def _representative_w_long(
    data: np.ndarray,
    labels: np.ndarray,
    excluded_indices: set[int],
    muscle_order: list[str],
    group_id: str,
) -> pd.DataFrame:
    """Average kept member vectors per cluster and L2-normalize the representative W."""
    rows: list[dict[str, Any]] = []
    for cluster_id in sorted(int(value) for value in np.unique(labels).tolist()):
        member_indices = [idx for idx, label in enumerate(labels.tolist()) if int(label) == cluster_id and idx not in excluded_indices]
        if not member_indices:
            continue
        representative = data[np.asarray(member_indices, dtype=int)].mean(axis=0, dtype=np.float64)
        norm = float(np.linalg.norm(representative))
        if norm > 0.0:
            representative = representative / norm
        for muscle, value in zip(muscle_order, representative.tolist(), strict=True):
            rows.append(
                {
                    "group_id": group_id,
                    "cluster_id": cluster_id,
                    "muscle": muscle,
                    "W_value": float(value),
                }
            )
    return pd.DataFrame(rows)


def _representative_h_long(
    h_long: pl.DataFrame,
    sample_map: list[dict[str, Any]],
    labels: np.ndarray,
    excluded_indices: set[int],
    group_id: str,
) -> pd.DataFrame:
    """Compute representative H (mean H per cluster) from minimal H long table."""
    mapping_rows = []
    for idx, (sample, cluster_id) in enumerate(zip(sample_map, labels.tolist())):
        if idx in excluded_indices:
            continue
        mapping_rows.append({
            "trial_id": sample["trial_id"],
            "component_index": sample["component_index"],
            "cluster_id": int(cluster_id),
        })
    mapping_df = pd.DataFrame(mapping_rows)

    h_group = h_long.filter(pl.col("group_id") == group_id).to_pandas()
    h_merged = h_group.merge(mapping_df, on=["trial_id", "component_index"], how="inner")

    rep_h = (
        h_merged.groupby(["cluster_id", "frame_idx"], as_index=False)["h_value"]
        .mean()
    )
    rep_h["group_id"] = group_id
    return rep_h[["group_id", "cluster_id", "frame_idx", "h_value"]]


def _cluster_member_counts(labels: np.ndarray, excluded_indices: set[int], n_clusters: int) -> pd.DataFrame:
    """Summarize cluster membership before and after duplicate-component exclusion."""
    rows = []
    for cluster_id in range(n_clusters):
        before = int(np.sum(labels == cluster_id))
        after = int(np.sum([(int(label) == cluster_id) and (idx not in excluded_indices) for idx, label in enumerate(labels.tolist())]))
        rows.append(
            {
                "cluster_id": cluster_id,
                "member_count_before_exclusion": before,
                "member_count_after_exclusion": after,
                "removed_component_count": before - after,
            }
        )
    return pd.DataFrame(rows)


def _write_csv(frame: pd.DataFrame, path: Path) -> None:
    """Write CSV outputs with BOM for Korean-safe spreadsheet viewing."""
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(path, index=False, encoding="utf-8-sig")


def _write_json(payload: dict[str, Any], path: Path) -> None:
    """Write JSON summaries with UTF-8 BOM to match repo defaults."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8-sig")


def _md5_file(path: Path) -> str:
    """Compute an MD5 checksum for validation manifests."""
    digest = hashlib.md5()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_checksum_manifest(outdir: Path) -> Path:
    """Record MD5 checksums for all generated analysis artifacts."""
    lines = []
    for path in sorted(outdir.rglob("*")):
        if not path.is_file() or path.name == "checksums.md5":
            continue
        lines.append(f"{_md5_file(path)}  {path.relative_to(outdir).as_posix()}")
    manifest_path = outdir / "checksums.md5"
    manifest_path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")
    return manifest_path


def _compare_reference_md5(new_figure_dir: Path, baseline_run: Path) -> pd.DataFrame:
    """Compare regenerated figure checksums against the baseline default_run figures."""
    comparisons = []
    for filename in (
        "cross_group_cosine_heatmap.png",
        "cross_group_decision_summary.png",
        "cross_group_matched_w.png",
    ):
        new_path = new_figure_dir / filename
        reference_path = baseline_run / "figures" / filename
        comparisons.append(
            {
                "file": filename,
                "new_md5": _md5_file(new_path),
                "reference_md5": _md5_file(reference_path),
                "matches_reference": _md5_file(new_path) == _md5_file(reference_path),
            }
        )
    return pd.DataFrame(comparisons)


def _dry_run_summary(
    step_features: list[TrialFeature],
    nonstep_rep_w: pd.DataFrame,
    metadata_summary: dict[str, Any],
    expected_duplicates: list[str],
) -> None:
    """Print concise reconstruction and expectation checks for dry-run mode."""
    print("=" * 72)
    print("Cosine Rerun Gap13 Duplicate Exclusion")
    print("=" * 72)
    print("[M1] Reconstruction summary")
    print(f"  Step trials reconstructed: {len(step_features)}")
    print(f"  Step components reconstructed: {sum(len(feature.component_ids) for feature in step_features)}")
    print(f"  Baseline nonstep representative clusters: {nonstep_rep_w['cluster_id'].nunique()}")
    print("[M2] Baseline K audit")
    print(
        "  Step baseline status: "
        f"gap raw K={metadata_summary['k_gap_raw']}, "
        f"selected K={metadata_summary['k_selected']}, "
        f"k_min_unique={metadata_summary['k_min_unique']}, "
        f"selection_status={metadata_summary['selection_status']}"
    )
    print(
        "  Expected duplicate trials at step K=13: "
        + (", ".join(expected_duplicates) if expected_duplicates else "none")
    )
    print("=" * 72)


def main() -> None:
    """Run the analysis-only rerun from baseline exports to new figures."""
    args = parse_args()
    cfg = _load_yaml(args.config)
    muscle_order = list(cfg["muscles"]["names"])
    threshold = float(cfg["cross_group_w_similarity"]["threshold"])
    clustering_cfg = dict(cfg["synergy_clustering"])
    baseline_run = args.baseline_run

    print("=" * 72)
    print("Cosine Rerun Gap13 Duplicate Exclusion")
    print("=" * 72)
    print("\n[M1] Loading baseline exports...")
    w_long = _read_csv(baseline_run / "all_minimal_units_W.csv")
    h_long = _read_csv(baseline_run / "all_minimal_units_H_long.csv")
    baseline_labels = _read_csv(baseline_run / "all_cluster_labels.csv")
    rep_w = _read_csv(baseline_run / "all_representative_W_posthoc.csv").to_pandas()
    metadata = _read_csv(baseline_run / "all_clustering_metadata.csv")

    rep_h_baseline = _read_csv(baseline_run / "all_representative_H_posthoc_long.csv")
    nonstep_rep_h = rep_h_baseline.filter(pl.col("group_id") == NONSTEP_GROUP_ID).to_pandas()

    step_features = _load_trial_features(w_long, h_long, STEP_GROUP_ID, muscle_order)
    step_data, step_sample_map = _stack_weight_vectors_from_baseline_order(
        w_long,
        baseline_labels,
        muscle_order,
    )
    nonstep_rep_w = rep_w.loc[rep_w["group_id"] == NONSTEP_GROUP_ID].copy()
    metadata_summary, observed_by_k, expected_duplicate_evidence = _parse_step_metadata(
        metadata,
        baseline_run,
        args.step_k,
    )
    expected_duplicates = sorted(expected_duplicate_evidence["trial_id"].drop_duplicates().tolist())

    if args.dry_run:
        _dry_run_summary(step_features, nonstep_rep_w, metadata_summary, expected_duplicates)
        print("\nDry run complete.")
        return

    _ensure_outdir(args.outdir, overwrite=args.overwrite)
    figure_dir = args.outdir / "figures"
    _configure_fonts()

    print("\n[M2] Rebuilding fixed-K step clustering and excluding duplicate components...")
    fit_seed = int(clustering_cfg.get("random_state", 42)) + (int(args.step_k) * 1000)
    observed_fit = _fit_best_kmeans_result(
        step_data,
        int(args.step_k),
        int(clustering_cfg.get("repeats", 25)),
        fit_seed,
        clustering_cfg,
    )
    observed_objective = float(observed_fit["objective"])
    expected_objective = float(observed_by_k[int(args.step_k)])
    objective_diff = abs(observed_objective - expected_objective)
    objective_matches = math.isclose(observed_objective, expected_objective, rel_tol=0.0, abs_tol=float(args.objective_atol))
    if not objective_matches:
        raise ValueError(
            f"Fixed-K objective mismatch at K={args.step_k}: "
            f"observed={observed_objective}, expected={expected_objective}, "
            f"abs_diff={objective_diff}, atol={args.objective_atol}"
        )

    min_duplicate_candidate = _select_min_duplicate_candidate(
        step_data,
        step_sample_map,
        int(args.step_k),
        clustering_cfg,
        observed_result=observed_fit,
    )
    step_labels = np.asarray(min_duplicate_candidate["labels"], dtype=np.int32)

    component_table, duplicate_summary, excluded_indices = _duplicate_component_decisions(
        step_data,
        step_labels,
        step_sample_map,
        int(args.step_k),
    )
    observed_duplicates = sorted(duplicate_summary["trial_id"].drop_duplicates().tolist())
    observed_duplicate_evidence = _normalize_duplicate_evidence(
        duplicate_summary,
        "component_indexes_in_duplicate_set",
    ).rename(columns={"component_indexes_in_duplicate_set": "component_indexes_json"})
    if not observed_duplicate_evidence.equals(expected_duplicate_evidence):
        raise ValueError(
            "Min-duplicate fixed-K candidate does not match baseline duplicate evidence. "
            f"observed={observed_duplicate_evidence.to_dict(orient='records')}, "
            f"expected={expected_duplicate_evidence.to_dict(orient='records')}"
        )

    step_rep_before = _representative_w_long(step_data, step_labels, set(), muscle_order, STEP_GROUP_ID)
    step_rep_after = _representative_w_long(step_data, step_labels, excluded_indices, muscle_order, STEP_GROUP_ID)
    member_counts = _cluster_member_counts(step_labels, excluded_indices, int(args.step_k))

    print(
        f"  Step K={args.step_k} objective check: observed={observed_objective:.6f}, "
        f"baseline={expected_objective:.6f}, abs_diff={objective_diff:.6f}, "
        f"within_atol={objective_matches}"
    )
    print(
        f"  Min-duplicate candidate objective: {min_duplicate_candidate['objective']:.6f} "
        f"(seed={min_duplicate_candidate['seed']}, restart={min_duplicate_candidate['restart_index']})"
    )
    print(f"  Duplicate trials in min-duplicate candidate: {', '.join(observed_duplicates)}")
    print(f"  Excluded duplicate components: {len(excluded_indices)}")

    print("\n[M3] Regenerating cross-group cosine artifacts and figures...")
    rerun_rep_w = pd.concat([step_rep_after, nonstep_rep_w], ignore_index=True)
    step_df, nonstep_df = build_cluster_w_matrix(rerun_rep_w, muscle_order)
    pairwise_df = compute_pairwise_cosine(step_df, nonstep_df)
    assigned_df = solve_assignment(pairwise_df)
    pairwise_output_df = annotate_pairwise_assignment(pairwise_df, assigned_df, threshold)
    decision_df = build_cluster_decision(step_df, nonstep_df, pairwise_df, assigned_df, threshold)

    save_cross_group_heatmap(
        pairwise_df=pairwise_output_df,
        threshold=threshold,
        cfg=cfg,
        output_path=figure_dir / "cross_group_cosine_heatmap.png",
    )
    save_cross_group_matched_w(
        step_df=step_df,
        nonstep_df=nonstep_df,
        decision_df=decision_df,
        muscle_names=muscle_order,
        cfg=cfg,
        output_path=figure_dir / "cross_group_matched_w.png",
    )
    step_rep_h = _representative_h_long(
        h_long, step_sample_map, step_labels, excluded_indices, STEP_GROUP_ID,
    )
    step_label_rows = []
    for idx, (sample, cid) in enumerate(zip(step_sample_map, step_labels.tolist())):
        if idx in excluded_indices:
            continue
        step_label_rows.append({
            "group_id": STEP_GROUP_ID,
            "trial_id": sample["trial_id"],
            "component_index": sample["component_index"],
            "cluster_id": int(cid),
        })
    nonstep_baseline_labels = baseline_labels.filter(
        pl.col("group_id") == NONSTEP_GROUP_ID
    ).select(["group_id", "trial_id", "component_index", "cluster_id"]).to_pandas()
    combined_labels = pd.concat(
        [pd.DataFrame(step_label_rows), nonstep_baseline_labels], ignore_index=True,
    )
    save_cross_group_matched_h(
        rep_h_step=step_rep_h,
        rep_h_nonstep=nonstep_rep_h,
        minimal_h=h_long.select(
            ["group_id", "trial_id", "component_index", "frame_idx", "h_value"]
        ).to_pandas(),
        labels=combined_labels,
        decision_df=decision_df,
        cfg=cfg,
        output_path=figure_dir / "cross_group_matched_h.png",
    )
    save_cross_group_decision_summary(
        decision_df=decision_df,
        threshold=threshold,
        cfg=cfg,
        output_path=figure_dir / "cross_group_decision_summary.png",
    )

    _write_csv(component_table, args.outdir / "step_k13_component_assignments.csv")
    _write_csv(duplicate_summary, args.outdir / "step_k13_duplicate_component_summary.csv")
    _write_csv(member_counts, args.outdir / "step_k13_cluster_member_counts.csv")
    _write_csv(step_rep_before, args.outdir / "step_k13_representative_W_before_exclusion.csv")
    _write_csv(step_rep_after, args.outdir / "step_k13_representative_W_after_exclusion.csv")
    _write_csv(pairwise_output_df, args.outdir / "cross_group_w_pairwise_cosine.csv")
    _write_csv(decision_df, args.outdir / "cross_group_w_cluster_decision.csv")

    md5_compare = _compare_reference_md5(figure_dir, baseline_run)
    _write_csv(md5_compare, args.outdir / "md5_compare_vs_default_run_figures.csv")

    summary = {
        "analysis_name": "cosine_rerun_gap13_duplicate_exclusion",
        "baseline_run": str(baseline_run),
        "config_path": str(args.config),
        "step_group_baseline_k_gap_raw": metadata_summary["k_gap_raw"],
        "step_group_baseline_k_selected": metadata_summary["k_selected"],
        "step_group_fixed_k_rerun": int(args.step_k),
        "expected_duplicate_trials_at_fixed_k": expected_duplicates,
        "observed_duplicate_trials_at_fixed_k": observed_duplicates,
        "excluded_duplicate_component_count": int(len(excluded_indices)),
        "step_cluster_count_after_exclusion": int(step_df["cluster_id"].nunique()),
        "nonstep_cluster_count_reference": int(nonstep_df["cluster_id"].nunique()),
        "same_synergy_pair_count": int(decision_df["match_id"].dropna().nunique()),
        "step_same_synergy_cluster_count": int(
            ((decision_df["group_id"] == STEP_GROUP_ID) & (decision_df["final_label"] == "same_synergy")).sum()
        ),
        "step_group_specific_cluster_count": int(
            ((decision_df["group_id"] == STEP_GROUP_ID) & (decision_df["final_label"] == "group_specific_synergy")).sum()
        ),
        "nonstep_same_synergy_cluster_count": int(
            ((decision_df["group_id"] == NONSTEP_GROUP_ID) & (decision_df["final_label"] == "same_synergy")).sum()
        ),
        "nonstep_group_specific_cluster_count": int(
            ((decision_df["group_id"] == NONSTEP_GROUP_ID) & (decision_df["final_label"] == "group_specific_synergy")).sum()
        ),
        "threshold": threshold,
        "observed_fixed_k_objective": observed_objective,
        "baseline_expected_fixed_k_objective": expected_objective,
        "observed_fixed_k_objective_abs_diff": objective_diff,
        "observed_fixed_k_objective_within_tolerance": bool(objective_matches),
        "objective_atol": float(args.objective_atol),
        "min_duplicate_candidate_objective": float(min_duplicate_candidate["objective"]),
        "min_duplicate_candidate_seed": int(min_duplicate_candidate["seed"]),
        "min_duplicate_candidate_restart_index": int(min_duplicate_candidate["restart_index"]),
        "reference_md5_all_match": bool(md5_compare["matches_reference"].all()),
    }
    _write_json(summary, args.outdir / "summary.json")
    checksum_path = _write_checksum_manifest(args.outdir)

    print("\nAnalysis complete.")
    print(f"  Output directory: {args.outdir}")
    print(f"  Checksums: {checksum_path}")
    print(
        "  Same-synergy pairs: "
        f"{summary['same_synergy_pair_count']} "
        f"(step same={summary['step_same_synergy_cluster_count']}, "
        f"step group-specific={summary['step_group_specific_cluster_count']})"
    )


if __name__ == "__main__":
    main()
