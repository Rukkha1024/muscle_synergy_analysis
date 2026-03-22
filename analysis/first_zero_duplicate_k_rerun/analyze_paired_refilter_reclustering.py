"""Run paired-only reclustering from one final parquet bundle.

This analysis keeps only paired `(subject, velocity)` concatenated
units, reruns first-zero-duplicate clustering, and exports paired
manifests, exact McNemar statistics, and a validated workbook.
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from pandas.api.types import is_bool_dtype, is_numeric_dtype
import polars as pl
from rich.console import Console

import analyze_first_zero_duplicate_k_rerun as legacy
from src.synergy_stats.single_parquet import load_single_parquet_bundle


DEFAULT_SOURCE_PARQUET = REPO_ROOT / "outputs" / "final_concatenated.parquet"
DEFAULT_CONFIG = REPO_ROOT / "configs" / "global_config.yaml"
DEFAULT_OUT_DIR = SCRIPT_DIR / "artifacts" / "paired_refilter_reclustering"
DEFAULT_GROUP_ID = "pooled_step_nonstep"
ERROR_TOKENS = {"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"}
WORKBOOK_SHEETS = ("summary", "cluster_stats", "paired_detail", "table_guide")
console = Console()


class TablePlacement:
    def __init__(self, table_name: str, sheet_name: str, start_row: int, start_col: int, frame: pd.DataFrame) -> None:
        self.table_name = table_name
        self.sheet_name = sheet_name
        self.start_row = start_row
        self.start_col = start_col
        self.frame = frame


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments for the paired rerun workflow."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-parquet", type=Path, default=DEFAULT_SOURCE_PARQUET)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--group-id", default=DEFAULT_GROUP_ID)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--algorithm", default=None)
    parser.add_argument("--max-clusters", type=int, default=None)
    parser.add_argument("--uniqueness-candidate-restarts", type=int, default=None)
    parser.add_argument("--max-iter", type=int, default=None)
    parser.add_argument("--random-state", type=int, default=None)
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def _bool_expr(column_name: str) -> pl.Expr:
    return pl.col(column_name).cast(pl.Float64, strict=False).fill_null(0.0) > 0.0


def _pair_key_expr() -> pl.Expr:
    return pl.concat_str(
        [
            pl.col("subject").cast(pl.Utf8),
            pl.lit("|"),
            pl.col("velocity").cast(pl.Float64, strict=False).cast(pl.Utf8),
        ]
    )


def _sheet_lines(description: str, key_columns: tuple[str, ...], example_lines: tuple[str, ...], notes: str = "") -> list[str]:
    lines = [f"[목적] {description}", "[핵심 컬럼]"]
    lines.extend(f"{index}. {line}" for index, line in enumerate(key_columns, start=1))
    lines.append("[예시]")
    lines.extend(example_lines)
    if notes:
        lines.append("[읽는 팁]")
        lines.append(notes)
    return lines


def _write_text_block(sheet, lines: list[str], start_row: int = 1, start_col: int = 1) -> int:
    row_idx = start_row
    for line in lines:
        cell = sheet.cell(row=row_idx, column=start_col, value=line)
        if line.startswith("["):
            cell.font = Font(bold=True)
        row_idx += 1
    return row_idx


def _infer_text_columns(frame: pd.DataFrame) -> set[str]:
    text_columns: set[str] = set()
    for column_name in frame.columns:
        series = frame[column_name]
        if is_bool_dtype(series) or not is_numeric_dtype(series):
            text_columns.add(column_name)
    return text_columns


def _coerce_cell_value(value: Any, *, force_text: bool) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if force_text:
        return str(value)
    return value


def _table_ref(start_row: int, start_col: int, frame: pd.DataFrame) -> str:
    end_row = start_row + len(frame.index)
    end_col = start_col + len(frame.columns) - 1
    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"


def _write_table(sheet, placement: TablePlacement) -> str:
    frame = placement.frame
    text_columns = _infer_text_columns(frame)
    for col_idx, column_name in enumerate(frame.columns, start=placement.start_col):
        header = sheet.cell(row=placement.start_row, column=col_idx, value=str(column_name))
        header.font = Font(bold=True)
    for row_offset, row in enumerate(frame.itertuples(index=False), start=1):
        for col_idx, (column_name, value) in enumerate(zip(frame.columns, row, strict=True), start=placement.start_col):
            force_text = column_name in text_columns
            cell = sheet.cell(
                row=placement.start_row + row_offset,
                column=col_idx,
                value=_coerce_cell_value(value, force_text=force_text),
            )
            if force_text:
                cell.number_format = "@"
    ref = _table_ref(placement.start_row, placement.start_col, frame)
    table = Table(displayName=placement.table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    for col_idx, column_name in enumerate(frame.columns, start=placement.start_col):
        width = max(len(str(column_name)), 14)
        for value in frame.iloc[:, col_idx - placement.start_col].astype(str).tolist():
            width = min(max(width, len(value) + 2), 60)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    return ref


def _table_guide_frame(table_rows: list[dict[str, str]]) -> pd.DataFrame:
    return pd.DataFrame(
        table_rows,
        columns=[
            "table_name",
            "sheet_name",
            "table_range",
            "description",
            "key_columns",
            "column_guide",
            "notes",
        ],
    )


def _load_group_frame(bundle: dict[str, Any], frame_key: str, group_id: str) -> pl.DataFrame:
    frame = bundle.get(frame_key)
    if frame is None or frame.empty:
        raise ValueError(f"Source parquet is missing required frame `{frame_key}`.")
    frame_pl = pl.from_pandas(frame)
    if "group_id" in frame_pl.columns:
        frame_pl = frame_pl.filter(pl.col("group_id").cast(pl.Utf8) == group_id)
    if frame_pl.is_empty():
        raise ValueError(f"Frame `{frame_key}` does not contain group `{group_id}`.")
    return frame_pl


def _build_pair_manifests(trial_windows_pl: pl.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, set[str]]:
    if trial_windows_pl.is_empty():
        raise ValueError("trial_windows is empty; cannot derive paired manifests.")
    prepared = trial_windows_pl.with_columns(
        _pair_key_expr().alias("paired_key"),
        _bool_expr("analysis_is_step").alias("analysis_is_step_bool"),
        _bool_expr("analysis_is_nonstep").alias("analysis_is_nonstep_bool"),
    )
    selection_column = None
    if "analysis_selected_group_prepaired" in prepared.columns:
        selection_column = "analysis_selected_group_prepaired"
    elif "analysis_selected_group" in prepared.columns:
        selection_column = "analysis_selected_group"
    if selection_column is not None:
        prepared = prepared.filter(_bool_expr(selection_column))
    summary = (
        prepared.group_by(["subject", "velocity", "paired_key"], maintain_order=True)
        .agg(
            pl.col("analysis_is_step_bool").any().alias("has_step"),
            pl.col("analysis_is_nonstep_bool").any().alias("has_nonstep"),
            pl.col("analysis_unit_id")
            .filter(pl.col("analysis_is_step_bool"))
            .drop_nulls()
            .first()
            .alias("step_analysis_unit_id"),
            pl.col("analysis_unit_id")
            .filter(pl.col("analysis_is_nonstep_bool"))
            .drop_nulls()
            .first()
            .alias("nonstep_analysis_unit_id"),
            pl.col("trial_id")
            .filter(pl.col("analysis_is_step_bool"))
            .drop_nulls()
            .first()
            .alias("step_trial_id"),
            pl.col("trial_id")
            .filter(pl.col("analysis_is_nonstep_bool"))
            .drop_nulls()
            .first()
            .alias("nonstep_trial_id"),
            pl.col("source_trial_nums_csv")
            .filter(pl.col("analysis_is_step_bool"))
            .drop_nulls()
            .first()
            .alias("step_source_trial_nums_csv"),
            pl.col("source_trial_nums_csv")
            .filter(pl.col("analysis_is_nonstep_bool"))
            .drop_nulls()
            .first()
            .alias("nonstep_source_trial_nums_csv"),
            pl.len().alias("analysis_unit_row_n"),
        )
        .with_columns(
            (pl.col("has_step") & pl.col("has_nonstep")).alias("analysis_is_paired_key"),
            pl.when(pl.col("has_step") & pl.col("has_nonstep"))
            .then(pl.lit("paired"))
            .when(pl.col("has_step"))
            .then(pl.lit("step_only"))
            .when(pl.col("has_nonstep"))
            .then(pl.lit("nonstep_only"))
            .otherwise(pl.lit("no_selected_rows"))
            .alias("analysis_pair_status"),
        )
    )
    paired_manifest = summary.filter(pl.col("analysis_is_paired_key"))
    excluded_manifest = summary.filter(~pl.col("analysis_is_paired_key"))
    included_analysis_units = set(
        paired_manifest.select(["step_analysis_unit_id", "nonstep_analysis_unit_id"])
        .to_pandas()
        .melt(value_name="analysis_unit_id")["analysis_unit_id"]
        .dropna()
        .astype(str)
        .tolist()
    )
    return paired_manifest.to_pandas(), excluded_manifest.to_pandas(), included_analysis_units


def _filter_frame_by_analysis_units(frame_pl: pl.DataFrame, analysis_unit_ids: set[str]) -> pd.DataFrame:
    filtered = frame_pl.filter(pl.col("analysis_unit_id").cast(pl.Utf8).is_in(sorted(analysis_unit_ids)))
    return filtered.to_pandas()


def _ensure_trial_window_counts(trial_windows_pl: pl.DataFrame, source_trial_windows_pl: pl.DataFrame) -> pl.DataFrame:
    if "analysis_source_trial_count" in trial_windows_pl.columns:
        return trial_windows_pl
    count_pl = source_trial_windows_pl.group_by("analysis_unit_id").agg(
        pl.len().alias("analysis_source_trial_count")
    )
    updated = trial_windows_pl.join(count_pl, on="analysis_unit_id", how="left")
    if "analysis_source_trial_count" not in updated.columns:
        return updated
    return updated.with_columns(
        pl.when(pl.col("analysis_source_trial_count").is_not_null())
        .then(pl.col("analysis_source_trial_count"))
        .when(pl.col("source_trial_nums_csv").is_not_null())
        .then(pl.col("source_trial_nums_csv").cast(pl.Utf8).str.split("|").list.len())
        .otherwise(pl.lit(0))
        .alias("analysis_source_trial_count")
    )


def _attach_trial_window_metadata(frame_pl: pl.DataFrame, trial_windows_pl: pl.DataFrame) -> pl.DataFrame:
    if "analysis_unit_id" not in trial_windows_pl.columns:
        return frame_pl
    metadata_pl = trial_windows_pl.unique(
        subset=["analysis_unit_id"],
        maintain_order=True,
    )
    join_columns = [
        column
        for column in metadata_pl.columns
        if column == "analysis_unit_id" or column not in frame_pl.columns
    ]
    if len(join_columns) <= 1:
        return frame_pl
    return frame_pl.join(metadata_pl.select(join_columns), on="analysis_unit_id", how="left")


def _bh_adjust(p_values: list[float]) -> list[float]:
    if not p_values:
        return []
    values = np.asarray(p_values, dtype=float)
    order = np.argsort(values)
    ranked = values[order]
    adjusted = np.empty_like(ranked)
    running = 1.0
    n_values = float(len(values))
    for index in range(len(ranked) - 1, -1, -1):
        rank = float(index + 1)
        running = min(running, float(ranked[index] * n_values / rank))
        adjusted[index] = min(running, 1.0)
    output = np.empty_like(adjusted)
    output[order] = adjusted
    return output.tolist()


def _exact_mcnemar_pvalue(step_only_n: int, nonstep_only_n: int) -> tuple[float, str]:
    discordant_n = int(step_only_n) + int(nonstep_only_n)
    if discordant_n == 0:
        return 1.0, "no_discordant_pairs"
    observed = int(step_only_n)
    observed_prob = math.comb(discordant_n, observed) / (2**discordant_n)
    p_value = 0.0
    for candidate in range(discordant_n + 1):
        candidate_prob = math.comb(discordant_n, candidate) / (2**discordant_n)
        if candidate_prob <= observed_prob + 1e-12:
            p_value += candidate_prob
    p_value = min(float(p_value), 1.0)
    return p_value, ""


def _presence_label(step_present: bool, nonstep_present: bool) -> str:
    if step_present and nonstep_present:
        return "both_present"
    if step_present:
        return "step_only"
    if nonstep_present:
        return "nonstep_only"
    return "both_absent"


def _build_presence_tables(
    paired_manifest: pd.DataFrame,
    labels_frame: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    labels_pdf = labels_frame.copy()
    labels_pdf["analysis_unit_id"] = labels_pdf["analysis_unit_id"].astype(str)
    labels_pdf["cluster_id"] = labels_pdf["cluster_id"].astype(int)
    presence_set = {
        (analysis_unit_id, int(cluster_id))
        for analysis_unit_id, cluster_id in labels_pdf[["analysis_unit_id", "cluster_id"]].drop_duplicates().itertuples(index=False)
    }
    cluster_ids = sorted(labels_pdf["cluster_id"].dropna().astype(int).unique().tolist())
    detail_rows: list[dict[str, Any]] = []
    for manifest_row in paired_manifest.itertuples(index=False):
        for cluster_id in cluster_ids:
            step_present = (str(manifest_row.step_analysis_unit_id), cluster_id) in presence_set
            nonstep_present = (str(manifest_row.nonstep_analysis_unit_id), cluster_id) in presence_set
            detail_rows.append(
                {
                    "cluster_id": int(cluster_id),
                    "subject": str(manifest_row.subject),
                    "velocity": float(manifest_row.velocity),
                    "paired_key": str(manifest_row.paired_key),
                    "step_present": bool(step_present),
                    "nonstep_present": bool(nonstep_present),
                    "presence_label": _presence_label(step_present, nonstep_present),
                }
            )
    detail_frame = pd.DataFrame(detail_rows)
    stats_rows: list[dict[str, Any]] = []
    q_values = _bh_adjust(
        [
            _exact_mcnemar_pvalue(
                int((cluster_frame["presence_label"] == "step_only").sum()),
                int((cluster_frame["presence_label"] == "nonstep_only").sum()),
            )[0]
            for _, cluster_frame in detail_frame.groupby("cluster_id", sort=True)
        ]
    )
    for q_value, (cluster_id, cluster_frame) in zip(q_values, detail_frame.groupby("cluster_id", sort=True), strict=True):
        paired_key_n = int(len(cluster_frame))
        step_present_n = int(cluster_frame["step_present"].sum())
        nonstep_present_n = int(cluster_frame["nonstep_present"].sum())
        both_present_n = int((cluster_frame["presence_label"] == "both_present").sum())
        step_only_n = int((cluster_frame["presence_label"] == "step_only").sum())
        nonstep_only_n = int((cluster_frame["presence_label"] == "nonstep_only").sum())
        both_absent_n = int((cluster_frame["presence_label"] == "both_absent").sum())
        mcnemar_p, mcnemar_note = _exact_mcnemar_pvalue(step_only_n, nonstep_only_n)
        step_presence_rate = float(step_present_n / paired_key_n) if paired_key_n else math.nan
        nonstep_presence_rate = float(nonstep_present_n / paired_key_n) if paired_key_n else math.nan
        diff = float(step_presence_rate - nonstep_presence_rate)
        if q_value < 0.05:
            interpretation_label = "strategy_biased"
        elif abs(diff) <= 0.15:
            interpretation_label = "shared_candidate"
        else:
            interpretation_label = "uncertain_not_significant"
        stats_rows.append(
            {
                "cluster_id": int(cluster_id),
                "paired_key_n": paired_key_n,
                "step_present_n": step_present_n,
                "nonstep_present_n": nonstep_present_n,
                "both_present_n": both_present_n,
                "step_only_n": step_only_n,
                "nonstep_only_n": nonstep_only_n,
                "both_absent_n": both_absent_n,
                "step_presence_rate": step_presence_rate,
                "nonstep_presence_rate": nonstep_presence_rate,
                "presence_rate_diff_step_minus_nonstep": diff,
                "mcnemar_p": mcnemar_p,
                "mcnemar_q_bh": float(q_value),
                "mcnemar_note": mcnemar_note,
                "interpretation_label": interpretation_label,
            }
        )
    return pd.DataFrame(stats_rows), detail_frame


def _write_csv(frame: pd.DataFrame, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def write_paired_cluster_statistics_workbook(
    path: Path,
    *,
    summary_frame: pd.DataFrame,
    cluster_stats_frame: pd.DataFrame,
    paired_detail_frame: pd.DataFrame,
) -> tuple[Path, dict[str, Any]]:
    """Write the paired statistics workbook and validate it after reopening."""
    workbook = Workbook()
    table_rows: list[dict[str, str]] = []

    sheet_specs = [
        (
            "summary",
            summary_frame,
            "tbl_paired_summary",
            "Paired rerun의 핵심 bookkeeping 값을 한 줄로 요약합니다.",
            (
                "paired_key_n: paired step/nonstep key 수입니다.",
                "analysis_unit_n_postpaired: paired gate 이후 rerun에 남은 analysis unit 수입니다.",
                "k_selected_first_zero_duplicate: rerun이 최종 선택한 first_zero_duplicate K입니다.",
            ),
            (
                "예를 들어 paired_key_n=2 이고 analysis_unit_n_postpaired=4 이면, 두 key가 step/nonstep 한 쌍씩 유지된 상태로 rerun이 수행된 것입니다.",
            ),
            "이 시트를 먼저 보고 cluster_stats와 paired_detail로 내려가면 읽기 쉽습니다.",
        ),
        (
            "cluster_stats",
            cluster_stats_frame,
            "tbl_paired_cluster_stats",
            "Cluster별 paired presence count와 exact McNemar 결과를 요약합니다.",
            (
                "step_only_n / nonstep_only_n: McNemar discordant pair count입니다.",
                "mcnemar_p / mcnemar_q_bh: exact McNemar raw p와 BH 보정 q입니다.",
                "interpretation_label: paired presence 차이에 대한 reviewer-facing 해석 라벨입니다.",
            ),
            (
                "예를 들어 q<0.05 이면 strategy_biased 로 읽고, q>=0.05 이면서 차이가 작으면 shared_candidate 로 해석합니다.",
            ),
            "step_only_n + nonstep_only_n 이 0이면 p=1.0, note=no_discordant_pairs 로 고정됩니다.",
        ),
        (
            "paired_detail",
            paired_detail_frame,
            "tbl_paired_cluster_detail",
            "Cluster x paired key 단위의 presence evidence를 한 줄씩 기록합니다.",
            (
                "paired_key: subject와 velocity를 묶은 paired identifier입니다.",
                "step_present / nonstep_present: 해당 cluster가 step/nonstep unit에 나타났는지 뜻합니다.",
                "presence_label: both_present, step_only, nonstep_only, both_absent 중 하나입니다.",
            ),
            (
                "예를 들어 step_present=True, nonstep_present=False 면 그 paired key는 해당 cluster에서 step_only row로 기록됩니다.",
            ),
            "cluster_stats count는 이 detail 시트의 presence_label 집계를 그대로 합친 결과여야 합니다.",
        ),
    ]

    for index, (sheet_name, frame, table_name, description, key_columns, example_lines, notes) in enumerate(sheet_specs):
        sheet = workbook.active if index == 0 else workbook.create_sheet(sheet_name)
        sheet.title = sheet_name
        start_row = _write_text_block(
            sheet,
            _sheet_lines(description, key_columns, example_lines, notes),
            start_row=1,
            start_col=1,
        ) + 1
        placement = TablePlacement(
            table_name=table_name,
            sheet_name=sheet_name,
            start_row=start_row,
            start_col=1,
            frame=frame,
        )
        ref = _write_table(sheet, placement)
        table_rows.append(
            {
                "table_name": table_name,
                "sheet_name": sheet_name,
                "table_range": ref,
                "description": description,
                "key_columns": ", ".join(column.split(":")[0] for column in key_columns),
                "column_guide": " | ".join(key_columns),
                "notes": notes,
            }
        )

    guide_sheet = workbook.create_sheet("table_guide")
    guide_lines = _sheet_lines(
        "Workbook 안의 모든 표와 읽는 순서를 요약합니다.",
        (
            "table_name: Excel Table 이름입니다.",
            "sheet_name: 표가 위치한 시트 이름입니다.",
            "column_guide: 주요 컬럼을 어떻게 읽는지 짧게 설명합니다.",
        ),
        (
            "예를 들어 tbl_paired_cluster_stats 행을 먼저 읽으면 cluster_stats 시트에서 어떤 열을 봐야 하는지 빠르게 찾을 수 있습니다.",
        ),
        "summary → cluster_stats → paired_detail 순서로 읽는 것을 권장합니다.",
    )
    guide_start_row = _write_text_block(guide_sheet, guide_lines, start_row=1, start_col=1) + 1
    guide_row = {
        "table_name": "tbl_table_guide",
        "sheet_name": "table_guide",
        "table_range": "pending",
        "description": "Workbook table inventory와 reading index입니다.",
        "key_columns": "table_name, sheet_name, table_range",
        "column_guide": "table_name: 표 이름 | sheet_name: 시트 이름 | table_range: 표 범위",
        "notes": "이 표 자체도 table_guide에 기록합니다.",
    }
    guide_frame = _table_guide_frame(table_rows + [guide_row])
    guide_ref = _write_table(
        guide_sheet,
        TablePlacement(
            table_name="tbl_table_guide",
            sheet_name="table_guide",
            start_row=guide_start_row,
            start_col=1,
            frame=guide_frame,
        ),
    )
    guide_range_col = guide_frame.columns.get_loc("table_range") + 1
    guide_self_row = guide_start_row + len(guide_frame.index)
    guide_sheet.cell(row=guide_self_row, column=guide_range_col, value=guide_ref).number_format = "@"

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    validation = validate_paired_cluster_statistics_workbook(path)
    return path, validation


def validate_paired_cluster_statistics_workbook(path: Path) -> dict[str, Any]:
    """Reopen the paired statistics workbook and validate sheet/table integrity."""
    expected_tables = {
        "summary": {"tbl_paired_summary"},
        "cluster_stats": {"tbl_paired_cluster_stats"},
        "paired_detail": {"tbl_paired_cluster_detail"},
        "table_guide": {"tbl_table_guide"},
    }
    issues = {"errors": [], "blanks": [], "tables": []}
    workbook = load_workbook(path)
    try:
        for sheet_name in WORKBOOK_SHEETS:
            if sheet_name not in workbook.sheetnames:
                issues["tables"].append((sheet_name, "missing_sheet"))
                continue
            sheet = workbook[sheet_name]
            actual_names = set(sheet.tables.keys())
            for table_name in sorted(expected_tables[sheet_name] - actual_names):
                issues["tables"].append((table_name, "missing_table"))
            guide_values = [sheet.cell(row=row_idx, column=1).value for row_idx in range(1, min(sheet.max_row, 24) + 1)]
            for label in ("[목적]", "[핵심 컬럼]", "[예시]"):
                if not any(isinstance(value, str) and value.startswith(label) for value in guide_values):
                    issues["blanks"].append((sheet_name, label))
            for table_name in actual_names:
                min_col, min_row, max_col, max_row = range_boundaries(sheet.tables[table_name].ref)
                for row in sheet.iter_rows(
                    min_row=min_row + 1,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                ):
                    for value in row:
                        if isinstance(value, str) and value in ERROR_TOKENS:
                            issues["errors"].append((sheet_name, table_name, value))

        if "table_guide" in workbook.sheetnames and "tbl_table_guide" in workbook["table_guide"].tables:
            guide_sheet = workbook["table_guide"]
            guide_table = guide_sheet.tables["tbl_table_guide"]
            min_col, min_row, max_col, max_row = range_boundaries(guide_table.ref)
            documented_tables: set[tuple[str, str]] = set()
            for row in guide_sheet.iter_rows(
                min_row=min_row + 1,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            ):
                if row[0] in (None, ""):
                    continue
                documented_tables.add((str(row[1]), str(row[0])))
                if row[3] in (None, "") or row[5] in (None, ""):
                    issues["blanks"].append(("table_guide", row[0]))
            actual_tables = {
                (sheet_name, table_name)
                for sheet_name in WORKBOOK_SHEETS
                if sheet_name in workbook.sheetnames
                for table_name in workbook[sheet_name].tables.keys()
            }
            for table_entry in sorted(actual_tables - documented_tables):
                issues["tables"].append((table_entry[1], "missing_guide_row"))
    finally:
        workbook.close()

    if any(issues.values()):
        raise ValueError(f"Paired statistics workbook validation failed for {path}: {issues}")
    return {
        "workbook_path": str(path),
        "sheet_count": len(WORKBOOK_SHEETS),
        "table_count": sum(len(table_names) for table_names in expected_tables.values()),
        "engine": "openpyxl",
        "excel_ui_visual_qa": "skipped_desktop_excel_unavailable",
        "fallback_reason": "openpyxl_default_wsl_headless_environment",
        "issues": issues,
    }


def main() -> None:
    args = parse_args()
    console.rule("[bold blue]Paired Refilter Reclustering[/bold blue]")
    console.print(f"Source parquet: {args.source_parquet}")
    console.print(f"Group: {args.group_id}")

    bundle = legacy._load_source_bundle(args.source_parquet)
    metadata_row = legacy._group_metadata_row(bundle, args.group_id)
    pipeline_duplicate_counts = legacy._parse_metric_json(metadata_row.get("duplicate_trial_count_by_k_json"))
    pipeline_k_gap_raw = legacy._scalar_to_int(metadata_row.get("k_gap_raw"))
    pipeline_k_selected = legacy._scalar_to_int(metadata_row.get("k_selected"))
    pipeline_k_min_unique = legacy._scalar_to_int(metadata_row.get("k_min_unique"))
    mode = legacy._mode_from_bundle(bundle, args.group_id)

    source_trial_windows_raw_pl = _load_group_frame(bundle, "source_trial_windows", args.group_id)
    trial_windows_pl = _ensure_trial_window_counts(
        _load_group_frame(bundle, "trial_windows", args.group_id),
        source_trial_windows_raw_pl,
    )
    paired_manifest, excluded_manifest, analysis_unit_ids = _build_pair_manifests(trial_windows_pl)
    if not analysis_unit_ids:
        raise ValueError("No paired analysis units were found in the source bundle.")
    console.print(
        f"Paired keys: {len(paired_manifest)} | Excluded keys: {len(excluded_manifest)} | "
        f"Analysis units after paired gate: {len(analysis_unit_ids)}"
    )
    if args.dry_run:
        console.print("Dry run complete. No artifacts were written.")
        return

    minimal_w_pl = _attach_trial_window_metadata(
        _load_group_frame(bundle, "minimal_W", args.group_id),
        trial_windows_pl,
    )
    minimal_h_pl = _attach_trial_window_metadata(
        _load_group_frame(bundle, "minimal_H_long", args.group_id),
        trial_windows_pl,
    )
    minimal_w = _filter_frame_by_analysis_units(minimal_w_pl, analysis_unit_ids)
    minimal_h = _filter_frame_by_analysis_units(minimal_h_pl, analysis_unit_ids)
    source_trial_windows_pl = _attach_trial_window_metadata(
        source_trial_windows_raw_pl,
        trial_windows_pl,
    )
    source_trial_windows_frame = _filter_frame_by_analysis_units(
        source_trial_windows_pl,
        analysis_unit_ids,
    )
    feature_rows, muscle_names = legacy._rebuild_feature_rows(minimal_w, minimal_h, args.group_id)
    scan_cfg = legacy._build_scan_cfg(metadata_row, args)
    inferred_k_max = legacy._default_k_max(metadata_row, sum(int(item.bundle.W_muscle.shape[1]) for item in feature_rows))
    resolved_k_max = args.max_clusters or inferred_k_max

    scan_result = legacy.scan_first_zero_duplicate_k(
        feature_rows,
        group_id=args.group_id,
        cfg=scan_cfg,
        k_max=resolved_k_max,
    )
    if scan_result["selected_k"] is None:
        raise ValueError("No zero-duplicate paired solution was found in the scanned K range.")
    console.print(f"First zero-duplicate K: {scan_result['selected_k']}")

    legacy._ensure_clean_dir(args.out_dir, overwrite=args.overwrite)
    cluster_result = legacy._cluster_result_from_scan(
        metadata_row=metadata_row,
        scan_result=scan_result,
        pipeline_duplicate_counts=pipeline_duplicate_counts,
    )
    export_cfg = legacy._export_cfg(
        base_cfg_path=args.config,
        out_dir=args.out_dir,
        mode=mode,
        muscle_names=muscle_names,
    )
    export_context = legacy._export_pipeline_like_outputs(
        out_dir=args.out_dir,
        mode=mode,
        cfg=export_cfg,
        feature_rows=feature_rows,
        cluster_result=cluster_result,
        group_id=args.group_id,
    )
    legacy._inject_source_trial_windows_and_rerender(
        bundle={"source_trial_windows": source_trial_windows_frame},
        out_dir=args.out_dir,
        cfg=export_cfg,
        mode=mode,
    )
    legacy._normalize_analysis_methods_manifest(args.out_dir)
    figure_path = legacy._plot_duplicate_burden(
        scan_result["scan_rows"],
        pipeline_duplicate_counts,
        pipeline_k_gap_raw,
        pipeline_k_selected,
        scan_result["selected_k"],
        args.out_dir / "k_duplicate_burden.png",
    )

    rerun_bundle = load_single_parquet_bundle(args.out_dir / f"final_{mode}.parquet")
    labels_frame = _load_group_frame(rerun_bundle, "labels", args.group_id).to_pandas()
    stats_frame, detail_frame = _build_presence_tables(paired_manifest, labels_frame)

    paired_manifest_path = _write_csv(paired_manifest, args.out_dir / "paired_subset_manifest.csv")
    excluded_manifest_path = _write_csv(excluded_manifest, args.out_dir / "excluded_nonpaired_manifest.csv")
    stats_path = _write_csv(stats_frame, args.out_dir / "paired_cluster_stats.csv")
    detail_path = _write_csv(detail_frame, args.out_dir / "paired_cluster_detail.csv")

    summary_frame = pd.DataFrame(
        [
            {
                "source_parquet": str(args.source_parquet.resolve()),
                "group_id": args.group_id,
                "paired_key_n": int(len(paired_manifest)),
                "excluded_pair_key_n": int(len(excluded_manifest)),
                "analysis_unit_n_postpaired": int(len(analysis_unit_ids)),
                "k_selected_first_zero_duplicate": int(scan_result["selected_k"]),
                "selection_method": "first_zero_duplicate",
                "resolved_mode": mode,
                "pipeline_k_gap_raw": pipeline_k_gap_raw,
                "pipeline_k_selected": pipeline_k_selected,
                "pipeline_k_min_unique": pipeline_k_min_unique,
            }
        ]
    )
    workbook_path, workbook_validation = write_paired_cluster_statistics_workbook(
        args.out_dir / "paired_cluster_statistics.xlsx",
        summary_frame=summary_frame,
        cluster_stats_frame=stats_frame,
        paired_detail_frame=detail_frame,
    )
    workbook_validation_summary = dict(workbook_validation)
    if isinstance(workbook_validation_summary.get("workbook_path"), str):
        workbook_validation_summary["workbook_path"] = legacy._path_relative_to(
            args.out_dir,
            workbook_validation_summary["workbook_path"],
        )

    k_scan_payload = {
        "selection_method": "first_zero_duplicate",
        "gap_statistic_used": False,
        "scan_rows": scan_result["scan_rows"],
    }
    summary_payload = {
        "source_parquet": str(args.source_parquet.resolve()),
        "group_id": args.group_id,
        "analysis_date": "2026-03-22",
        "resolved_mode": mode,
        "vector_count": scan_result["vector_count"],
        "trial_count": scan_result["trial_count"],
        "muscle_count": len(muscle_names),
        "k_min": scan_result["k_min"],
        "k_max": scan_result["k_max"],
        "selection_method": "first_zero_duplicate",
        "gap_statistic_used": False,
        "k_selected_first_zero_duplicate": int(scan_result["selected_k"]),
        "pipeline_k_gap_raw": pipeline_k_gap_raw,
        "pipeline_k_selected": pipeline_k_selected,
        "pipeline_k_min_unique": pipeline_k_min_unique,
        "paired_key_n": int(len(paired_manifest)),
        "excluded_pair_key_n": int(len(excluded_manifest)),
        "analysis_unit_n_postpaired": int(len(analysis_unit_ids)),
        "paired_subset_manifest_path": legacy._path_relative_to(args.out_dir, str(paired_manifest_path)),
        "excluded_nonpaired_manifest_path": legacy._path_relative_to(args.out_dir, str(excluded_manifest_path)),
        "paired_cluster_stats_csv_path": legacy._path_relative_to(args.out_dir, str(stats_path)),
        "paired_cluster_detail_csv_path": legacy._path_relative_to(args.out_dir, str(detail_path)),
        "paired_cluster_statistics_workbook_path": legacy._path_relative_to(args.out_dir, str(workbook_path)),
        "paired_cluster_statistics_workbook_validation": workbook_validation_summary,
        "duplicate_trial_count_by_k": {str(row["k"]): int(row["duplicate_trial_count"]) for row in scan_result["scan_rows"]},
        "pipeline_duplicate_trial_count_by_k": {str(key): int(value) for key, value in pipeline_duplicate_counts.items()},
        "scan_cfg": scan_cfg,
        "figure_path": figure_path.name,
        "final_parquet_path": legacy._path_relative_to(args.out_dir, export_context["artifacts"]["final_parquet_path"]),
        "final_parquet_alias_paths": {
            key: legacy._path_relative_to(args.out_dir, value)
            for key, value in export_context["artifacts"]["final_parquet_alias_paths"].items()
        },
        "mode_output_dirs": {
            key: legacy._path_relative_to(args.out_dir, value)
            for key, value in export_context["artifacts"]["mode_output_dirs"].items()
        },
    }
    summary_path = args.out_dir / "summary.json"
    k_scan_path = args.out_dir / "k_scan.json"
    legacy._write_json(summary_path, summary_payload)
    legacy._write_json(k_scan_path, k_scan_payload)

    checksum_targets = [
        path
        for path in args.out_dir.rglob("*")
        if path.is_file() and path.name != "checksums.md5"
    ]
    legacy._write_checksums(checksum_targets, args.out_dir / "checksums.md5")

    console.print(f"Paired workbook: {workbook_path}")
    console.print(f"Summary: {summary_path}")


if __name__ == "__main__":
    main()
