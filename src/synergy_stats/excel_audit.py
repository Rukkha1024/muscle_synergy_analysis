"""Build and validate the clustering audit workbook.

This helper converts clustering diagnostics into Excel tables,
writes a compact workbook with summary guidance and examples,
and reopens the saved file to verify the expected sheets
and table inventory are present.
"""

from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import polars as pl


GUIDE_LINES = [
    "[목적] 이 워크북은 각 clustering K가 왜 채택되었는지와 trial 내부 중복이 어디서 발생했는지를 설명합니다.",
    "[표 읽는 순서]",
    "1. k_gap_raw 는 gap statistic이 처음 선택한 K입니다.",
    "2. k_min_unique 는 trial 내부 중복이 0개가 되는 첫 번째 K입니다.",
    "3. k_selected 는 이번 실행에서 최종 채택된 K입니다.",
    "4. duplicate_trial_count 는 해당 K에서 아직 중복 cluster 할당이 남아 있는 trial 수입니다.",
    "5. audit table에서는 is_gap_raw_k, is_selected_k, is_first_zero_duplicate_k 를 먼저 보면 핵심 K를 빠르게 찾을 수 있습니다.",
    "[예시]",
    "예를 들어 k_gap_raw=6, k_min_unique=7, k_selected=7 이면, gap statistic이 먼저 고른 K=6에는 중복 trial이 남아 있어서 중복이 처음 0개가 되는 K=7을 최종 채택했다는 뜻입니다.",
    "어떤 trial에서 중복이 발생했는지는 tbl_duplicate_trial_summary 에서 보고, 그 중복이 어떤 cluster에서 나왔는지는 tbl_duplicate_cluster_detail 에서 확인하면 됩니다.",
    "[실행 환경 메모]",
    "Workbook 엔진: openpyxl 기본값.",
    "사유: WSL 또는 headless 환경에서는 openpyxl을 기본 workbook 엔진으로 사용합니다.",
    "Excel UI 시각 검수: 데스크톱 Excel 자동화를 사용할 수 없어 건너뛰었습니다.",
]

DUPLICATES_GUIDE_LINES = [
    "[목적] 이 시트는 후보 K에서 실제로 중복 cluster 할당이 발생한 trial과 세부 cluster 충돌 내역을 보여줍니다.",
    "[표 읽는 순서]",
    "1. 먼저 tbl_duplicate_trial_summary 에서 어떤 group_id, k, trial_id 에 중복이 있었는지 확인합니다.",
    "2. duplicate_cluster_count 는 해당 trial 안에서 몇 개의 cluster가 중복되었는지를 뜻합니다.",
    "3. duplicate_component_indexes_json 은 중복에 관련된 component index 목록입니다.",
    "4. 더 자세한 충돌 원인은 아래 tbl_duplicate_cluster_detail 에서 cluster_id 와 component_indexes_json 으로 확인합니다.",
    "[예시]",
    "예를 들어 trial_id=S01_v1_T2, k=9, duplicate_cluster_count=2 라면, K=9에서 이 trial 안에 서로 다른 component가 같은 cluster로 두 번 이상 묶였다는 뜻입니다.",
    "이때 아래 표에서 같은 trial_id 의 cluster_id 와 component_indexes_json 을 보면 어떤 cluster에서 중복이 났는지 바로 확인할 수 있습니다.",
]

TABLE_GUIDE_LINES = [
    "[목적] 이 시트는 워크북 안에 있는 모든 Excel Table의 위치와 해석 방법을 한눈에 정리한 안내표입니다.",
    "[표 읽는 순서]",
    "1. table_name 은 실제 Excel Table 이름입니다.",
    "2. sheet_name 과 table_range 를 함께 보면 해당 표가 워크북 어디에 있는지 찾을 수 있습니다.",
    "3. description 은 표가 무엇을 요약하는지 설명합니다.",
    "4. key_columns 는 해석할 때 먼저 봐야 하는 핵심 컬럼입니다.",
    "5. notes 는 읽을 때 주의할 점이나 바로 적용할 해석 팁입니다.",
    "[예시]",
    "예를 들어 tbl_duplicate_trial_summary 행을 보면 duplicates 시트의 어느 범위에 trial 중복 요약표가 있는지 찾을 수 있고, key_columns 를 보고 group_id, k, trial_id 를 먼저 확인하면 됩니다.",
]


ERROR_TOKENS = {"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"}


@dataclass(frozen=True)
class TablePlacement:
    table_name: str
    sheet_name: str
    start_row: int
    start_col: int
    frame: pd.DataFrame
    description: str
    key_columns: str
    notes: str = ""


def _json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _scalar_or_blank(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return value


def _flag_summary_text(group_id: str, k_gap_raw: int, k_selected: int, duplicate_count_at_gap_raw: int) -> str:
    if duplicate_count_at_gap_raw == 0 and k_gap_raw == k_selected:
        return (
            f"{group_id}: gap raw K={k_gap_raw} had zero duplicate trials, "
            f"so selected K={k_selected}."
        )
    if duplicate_count_at_gap_raw == 0:
        return (
            f"{group_id}: gap raw K={k_gap_raw} was already duplicate-free, "
            f"but selected K={k_selected} was exported by the current selection path."
        )
    return (
        f"{group_id}: gap raw K={k_gap_raw} had {duplicate_count_at_gap_raw} duplicate trial(s), "
        f"so the first zero-duplicate K={k_selected} was selected."
    )


def build_audit_tables(cluster_group_results: dict[str, dict[str, Any]]) -> dict[str, pd.DataFrame]:
    """Return workbook-ready tables from clustering results."""
    selection_rows: list[dict[str, Any]] = []
    k_rows: list[dict[str, Any]] = []
    duplicate_trial_rows: list[dict[str, Any]] = []
    duplicate_cluster_rows: list[dict[str, Any]] = []

    for group_id, payload in cluster_group_results.items():
        cluster_result = payload["cluster_result"]
        payload_group_id = str(payload.get("group_id", group_id))
        aggregation_mode = payload.get("aggregation_mode", "")
        summary_group_id = (
            f"{aggregation_mode}/{payload_group_id}"
            if aggregation_mode
            else payload_group_id
        )
        gap_by_k = {int(key): float(value) for key, value in cluster_result.get("gap_by_k", {}).items()}
        gap_sd_by_k = {int(key): float(value) for key, value in cluster_result.get("gap_sd_by_k", {}).items()}
        observed_objective_by_k = {
            int(key): float(value) for key, value in cluster_result.get("observed_objective_by_k", {}).items()
        }
        feasible_objective_by_k = cluster_result.get("feasible_objective_by_k", {})
        duplicate_trial_count_by_k = {
            int(key): int(value) for key, value in cluster_result.get("duplicate_trial_count_by_k", {}).items()
        }
        duplicate_trial_evidence_by_k = cluster_result.get("duplicate_trial_evidence_by_k", {})
        k_gap_raw = int(cluster_result.get("k_gap_raw", 0))
        k_selected = int(cluster_result.get("k_selected", 0))
        k_min_unique_raw = cluster_result.get("k_min_unique", "")
        k_min_unique = (
            int(k_min_unique_raw)
            if k_min_unique_raw not in ("", None) and not pd.isna(k_min_unique_raw)
            else ""
        )
        duplicate_count_at_gap_raw = int(duplicate_trial_count_by_k.get(k_gap_raw, 0))
        duplicate_count_at_selected = int(duplicate_trial_count_by_k.get(k_selected, 0))
        selection_row = {
            "group_id": payload_group_id,
            "selection_status": cluster_result.get("selection_status", ""),
            "k_gap_raw": k_gap_raw,
            "k_selected": k_selected,
            "k_min_unique": k_min_unique,
            "duplicate_trial_count_at_gap_raw": duplicate_count_at_gap_raw,
            "duplicate_trial_count_at_selected_k": duplicate_count_at_selected,
            "summary_text": _flag_summary_text(
                summary_group_id,
                k_gap_raw,
                k_selected,
                duplicate_count_at_gap_raw,
            ),
        }
        if aggregation_mode:
            selection_row["aggregation_mode"] = aggregation_mode
        selection_rows.append(selection_row)

        for k in sorted(gap_by_k):
            k_row = {
                "group_id": payload_group_id,
                "k": int(k),
                "gap": gap_by_k.get(k, ""),
                "gap_sd": gap_sd_by_k.get(k, ""),
                "observed_objective": observed_objective_by_k.get(k, ""),
                "feasible_objective": _scalar_or_blank(feasible_objective_by_k.get(k, "")),
                "duplicate_trial_count": duplicate_trial_count_by_k.get(k, 0),
                "is_gap_raw_k": k == k_gap_raw,
                "is_selected_k": k == k_selected,
                "is_first_zero_duplicate_k": k_min_unique != "" and k == k_min_unique,
            }
            if aggregation_mode:
                k_row["aggregation_mode"] = aggregation_mode
            k_rows.append(k_row)
            for trial_row in duplicate_trial_evidence_by_k.get(k, []):
                duplicate_trial_row = {
                    "group_id": payload_group_id,
                    "k": int(k),
                    "subject": trial_row["subject"],
                    "velocity": trial_row["velocity"],
                    "trial_num": trial_row["trial_num"],
                    "trial_id": trial_row["trial_id"],
                    "n_synergies_in_trial": int(trial_row["n_synergies_in_trial"]),
                    "duplicate_cluster_labels_json": _json_text(trial_row["duplicate_cluster_labels"]),
                    "duplicate_component_indexes_json": _json_text(trial_row["duplicate_component_indexes"]),
                    "duplicate_cluster_count": int(trial_row["duplicate_cluster_count"]),
                    "duplicate_component_count": int(trial_row["duplicate_component_count"]),
                    "is_gap_raw_k": k == k_gap_raw,
                    "is_selected_k": k == k_selected,
                }
                if aggregation_mode:
                    duplicate_trial_row["aggregation_mode"] = aggregation_mode
                duplicate_trial_rows.append(duplicate_trial_row)
                for cluster_detail in trial_row["duplicate_cluster_details"]:
                    duplicate_cluster_row = {
                        "group_id": payload_group_id,
                        "k": int(k),
                        "subject": trial_row["subject"],
                        "velocity": trial_row["velocity"],
                        "trial_num": trial_row["trial_num"],
                        "trial_id": trial_row["trial_id"],
                        "cluster_id": int(cluster_detail["cluster_id"]),
                        "component_indexes_json": _json_text(cluster_detail["component_indexes"]),
                        "component_count": int(cluster_detail["component_count"]),
                        "is_gap_raw_k": k == k_gap_raw,
                        "is_selected_k": k == k_selected,
                    }
                    if aggregation_mode:
                        duplicate_cluster_row["aggregation_mode"] = aggregation_mode
                    duplicate_cluster_rows.append(duplicate_cluster_row)

    selection_frame = _pl_to_pandas(pl.DataFrame(selection_rows))
    k_audit_frame = _pl_to_pandas(pl.DataFrame(k_rows))
    duplicate_trial_frame = _pl_to_pandas(pl.DataFrame(duplicate_trial_rows)) if duplicate_trial_rows else pd.DataFrame()
    duplicate_cluster_frame = _pl_to_pandas(pl.DataFrame(duplicate_cluster_rows)) if duplicate_cluster_rows else pd.DataFrame()
    return {
        "selection_summary": selection_frame,
        "k_audit": k_audit_frame,
        "duplicate_trial_summary": duplicate_trial_frame,
        "duplicate_cluster_detail": duplicate_cluster_frame,
    }


def _placeholder_frame(columns: list[str], message_column: str, message: str) -> pd.DataFrame:
    row = {column: "" for column in columns}
    row[message_column] = message
    return pd.DataFrame([row], columns=columns)


def _pl_to_pandas(frame: pl.DataFrame) -> pd.DataFrame:
    return pd.DataFrame(frame.to_dicts(), columns=frame.columns)


def _write_text_block(sheet, lines: list[str], start_row: int = 1, start_col: int = 1) -> int:
    row = start_row
    for line in lines:
        cell = sheet.cell(row=row, column=start_col, value=line)
        if line.startswith("["):
            cell.font = Font(bold=True)
        row += 1
    return row


def _cell_text(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return value


def _table_ref(start_row: int, start_col: int, frame: pd.DataFrame) -> str:
    end_row = start_row + len(frame.index)
    end_col = start_col + len(frame.columns) - 1
    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"


def _write_table(sheet, placement: TablePlacement, *, text_columns: set[str] | None = None) -> tuple[str, int]:
    text_columns = text_columns or set()
    frame = placement.frame
    for offset, column_name in enumerate(frame.columns, start=placement.start_col):
        header_cell = sheet.cell(row=placement.start_row, column=offset, value=str(column_name))
        header_cell.font = Font(bold=True)
    for row_offset, row in enumerate(frame.itertuples(index=False), start=1):
        for col_offset, (column_name, value) in enumerate(zip(frame.columns, row, strict=True), start=placement.start_col):
            cell = sheet.cell(row=placement.start_row + row_offset, column=col_offset, value=_cell_text(value))
            if column_name in text_columns:
                cell.number_format = "@"
    ref = _table_ref(placement.start_row, placement.start_col, frame)
    table = Table(displayName=placement.table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    for index, column_name in enumerate(frame.columns, start=placement.start_col):
        width = max(len(str(column_name)), 14)
        for value in frame.iloc[:, index - placement.start_col].astype(str).tolist():
            width = min(max(width, len(value) + 2), 60)
        sheet.column_dimensions[get_column_letter(index)].width = width
    return ref, placement.start_row + len(frame.index) + 3


def _table_guide_rows(table_info: list[dict[str, str]]) -> pd.DataFrame:
    return pd.DataFrame(
        table_info,
        columns=["table_name", "sheet_name", "table_range", "description", "key_columns", "notes"],
    )


def write_clustering_audit_workbook(path: Path, cluster_group_results: dict[str, dict[str, Any]]) -> Path:
    """Write the audit workbook and validate it after reopening."""
    tables = build_audit_tables(cluster_group_results)
    selection_frame = tables["selection_summary"]
    k_audit_frame = tables["k_audit"]
    duplicate_trial_frame = tables["duplicate_trial_summary"]
    duplicate_cluster_frame = tables["duplicate_cluster_detail"]

    include_aggregation_mode = "aggregation_mode" in selection_frame.columns
    duplicate_trial_columns = [
        * (["aggregation_mode"] if include_aggregation_mode else []),
        "group_id",
        "k",
        "subject",
        "velocity",
        "trial_num",
        "trial_id",
        "n_synergies_in_trial",
        "duplicate_cluster_labels_json",
        "duplicate_component_indexes_json",
        "duplicate_cluster_count",
        "duplicate_component_count",
        "is_gap_raw_k",
        "is_selected_k",
    ]
    duplicate_cluster_columns = [
        * (["aggregation_mode"] if include_aggregation_mode else []),
        "group_id",
        "k",
        "subject",
        "velocity",
        "trial_num",
        "trial_id",
        "cluster_id",
        "component_indexes_json",
        "component_count",
        "is_gap_raw_k",
        "is_selected_k",
    ]
    if duplicate_trial_frame.empty:
        duplicate_trial_frame = _placeholder_frame(
            duplicate_trial_columns,
            "trial_id",
            "No duplicate trials were found across the audited K range.",
        )
    if duplicate_cluster_frame.empty:
        duplicate_cluster_frame = _placeholder_frame(
            duplicate_cluster_columns,
            "trial_id",
            "No duplicate cluster details were found across the audited K range.",
        )

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    duplicates_sheet = workbook.create_sheet("duplicates")
    guide_sheet = workbook.create_sheet("table_guide")

    next_row = _write_text_block(summary_sheet, GUIDE_LINES, start_row=1, start_col=1) + 1
    table_info: list[dict[str, str]] = []
    summary_placements = [
        TablePlacement(
            table_name="tbl_clustering_selection_summary",
            sheet_name="summary",
            start_row=next_row,
            start_col=1,
            frame=selection_frame,
            description="One row per group showing the selected K and the short selection rationale.",
            key_columns="group_id, k_gap_raw, k_selected",
            notes="Read this table first.",
        ),
        TablePlacement(
            table_name="tbl_clustering_k_audit",
            sheet_name="summary",
            start_row=0,
            start_col=1,
            frame=k_audit_frame,
            description="One row per group and candidate K showing gap and duplicate burden.",
            key_columns="group_id, k",
            notes="Use the boolean flags to find the key K values.",
        ),
    ]
    ref, next_row = _write_table(
        summary_sheet,
        summary_placements[0],
        text_columns={"aggregation_mode", "group_id", "selection_status", "summary_text"},
    )
    table_info.append(
        {
            "table_name": summary_placements[0].table_name,
            "sheet_name": summary_placements[0].sheet_name,
            "table_range": ref,
            "description": summary_placements[0].description,
            "key_columns": summary_placements[0].key_columns,
            "notes": summary_placements[0].notes,
        }
    )
    summary_placements[1] = TablePlacement(
        table_name=summary_placements[1].table_name,
        sheet_name=summary_placements[1].sheet_name,
        start_row=next_row,
        start_col=summary_placements[1].start_col,
        frame=summary_placements[1].frame,
        description=summary_placements[1].description,
        key_columns=summary_placements[1].key_columns,
        notes=summary_placements[1].notes,
    )
    ref, _ = _write_table(
        summary_sheet,
        summary_placements[1],
        text_columns={"aggregation_mode", "group_id"},
    )
    table_info.append(
        {
            "table_name": summary_placements[1].table_name,
            "sheet_name": summary_placements[1].sheet_name,
            "table_range": ref,
            "description": summary_placements[1].description,
            "key_columns": summary_placements[1].key_columns,
            "notes": summary_placements[1].notes,
        }
    )

    duplicates_start_row = _write_text_block(duplicates_sheet, DUPLICATES_GUIDE_LINES, start_row=1, start_col=1) + 1
    duplicates_placements = [
        TablePlacement(
            table_name="tbl_duplicate_trial_summary",
            sheet_name="duplicates",
            start_row=duplicates_start_row,
            start_col=1,
            frame=duplicate_trial_frame,
            description="One row per duplicate trial at a candidate K.",
            key_columns="group_id, k, trial_id",
            notes="Placeholder row means no duplicate trials were observed.",
        ),
        TablePlacement(
            table_name="tbl_duplicate_cluster_detail",
            sheet_name="duplicates",
            start_row=0,
            start_col=1,
            frame=duplicate_cluster_frame,
            description="One row per duplicate cluster within a duplicate trial.",
            key_columns="group_id, k, trial_id, cluster_id",
            notes="Use this to inspect which component indexes collided.",
        ),
    ]
    ref, next_row = _write_table(
        duplicates_sheet,
        duplicates_placements[0],
        text_columns={
            "aggregation_mode",
            "group_id",
            "subject",
            "velocity",
            "trial_num",
            "trial_id",
            "duplicate_cluster_labels_json",
            "duplicate_component_indexes_json",
        },
    )
    table_info.append(
        {
            "table_name": duplicates_placements[0].table_name,
            "sheet_name": duplicates_placements[0].sheet_name,
            "table_range": ref,
            "description": duplicates_placements[0].description,
            "key_columns": duplicates_placements[0].key_columns,
            "notes": duplicates_placements[0].notes,
        }
    )
    duplicates_placements[1] = TablePlacement(
        table_name=duplicates_placements[1].table_name,
        sheet_name=duplicates_placements[1].sheet_name,
        start_row=next_row,
        start_col=duplicates_placements[1].start_col,
        frame=duplicates_placements[1].frame,
        description=duplicates_placements[1].description,
        key_columns=duplicates_placements[1].key_columns,
        notes=duplicates_placements[1].notes,
    )
    ref, _ = _write_table(
        duplicates_sheet,
        duplicates_placements[1],
        text_columns={
            "aggregation_mode",
            "group_id",
            "subject",
            "velocity",
            "trial_num",
            "trial_id",
            "component_indexes_json",
        },
    )
    table_info.append(
        {
            "table_name": duplicates_placements[1].table_name,
            "sheet_name": duplicates_placements[1].sheet_name,
            "table_range": ref,
            "description": duplicates_placements[1].description,
            "key_columns": duplicates_placements[1].key_columns,
            "notes": duplicates_placements[1].notes,
        }
    )

    guide_start_row = _write_text_block(guide_sheet, TABLE_GUIDE_LINES, start_row=1, start_col=1) + 1
    guide_frame = _table_guide_rows(table_info)
    guide_row = {
        "table_name": "tbl_table_guide",
        "sheet_name": "table_guide",
        "table_range": "pending",
        "description": "Inventory of workbook tables and how to interpret them.",
        "key_columns": "table_name, sheet_name",
        "notes": "Every workbook table must appear here.",
    }
    guide_frame = _table_guide_rows(table_info + [guide_row])
    guide_placement = TablePlacement(
        table_name="tbl_table_guide",
        sheet_name="table_guide",
        start_row=guide_start_row,
        start_col=1,
        frame=guide_frame,
        description="Inventory of workbook tables and how to interpret them.",
        key_columns="table_name, sheet_name",
        notes="Every workbook table must appear here.",
    )
    guide_ref, _ = _write_table(
        guide_sheet,
        guide_placement,
        text_columns={"table_name", "sheet_name", "table_range", "description", "key_columns", "notes"},
    )
    guide_table_range_col = guide_frame.columns.get_loc("table_range") + 1
    guide_self_row = guide_placement.start_row + len(guide_frame.index)
    guide_sheet.cell(row=guide_self_row, column=guide_table_range_col, value=guide_ref).number_format = "@"

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    validate_clustering_audit_workbook(path)
    return path


def validate_clustering_audit_workbook(path: Path) -> dict[str, Any]:
    """Reopen the workbook and confirm the expected sheets and tables exist."""
    expected_sheets = {"summary", "duplicates", "table_guide"}
    expected_tables = {
        "summary": {"tbl_clustering_selection_summary", "tbl_clustering_k_audit"},
        "duplicates": {"tbl_duplicate_trial_summary", "tbl_duplicate_cluster_detail"},
        "table_guide": {"tbl_table_guide"},
    }
    issues = {"errors": [], "blanks": [], "tables": []}
    workbook = load_workbook(path)
    try:
        sheet_names = set(workbook.sheetnames)
        missing_sheets = sorted(expected_sheets - sheet_names)
        for sheet_name in missing_sheets:
            issues["tables"].append((sheet_name, "missing_sheet"))

        actual_workbook_tables: set[tuple[str, str]] = set()
        for sheet_name, table_names in expected_tables.items():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            actual_table_names = set(sheet.tables.keys())
            actual_workbook_tables.update((sheet_name, table_name) for table_name in actual_table_names)
            for table_name in sorted(table_names - actual_table_names):
                issues["tables"].append((table_name, "missing_table"))

        required_guide_cells = {
            "summary": ("A1", "A2", "A8", "A9", "A11", "A12", "A13", "A14"),
            "duplicates": ("A1", "A2", "A7", "A8", "A9"),
            "table_guide": ("A1", "A2", "A8", "A9"),
        }
        for sheet_name, coords in required_guide_cells.items():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            for coord in coords:
                value = sheet[coord].value
                if value in (None, ""):
                    issues["blanks"].append((sheet_name, coord))

        guide_sheet = workbook["table_guide"] if "table_guide" in workbook.sheetnames else None
        guide_table = None if guide_sheet is None else guide_sheet.tables.get("tbl_table_guide")
        if guide_sheet is None:
            pass
        elif guide_table is None:
            issues["tables"].append(("tbl_table_guide", "missing_table"))
        else:
            min_col, min_row, max_col, max_row = range_boundaries(guide_table.ref)
            rows = list(
                guide_sheet.iter_rows(
                    min_row=min_row + 1,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
            )
            documented_tables: set[tuple[str, str]] = set()
            for row in rows:
                if row[0] in (None, ""):
                    continue
                documented_tables.add((str(row[1]), str(row[0])))
                if any(cell in ERROR_TOKENS for cell in row if isinstance(cell, str)):
                    issues["errors"].append(("table_guide", row[0], "error_token"))
                if row[3] in (None, ""):
                    issues["blanks"].append(("table_guide", row[0]))
            for table_entry in sorted(actual_workbook_tables - documented_tables):
                issues["tables"].append((table_entry[1], "missing_guide_row"))
    finally:
        workbook.close()

    if any(issues.values()):
        raise ValueError(f"Workbook validation failed for {path}: {issues}")
    return {
        "workbook_path": str(path),
        "sheet_count": 3,
        "table_count": sum(len(values) for values in expected_tables.values()),
        "engine": "openpyxl",
        "excel_ui_visual_qa": "skipped_desktop_excel_unavailable",
        "fallback_reason": "openpyxl_default_wsl_headless_environment",
        "issues": issues,
    }
