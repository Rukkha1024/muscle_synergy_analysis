"""Build a reader-friendly Excel workbook for run-level outputs.

This module writes one table per sheet with Korean guidance,
adds a `table_guide` inventory,
and validates the saved workbook after reopening it.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from pandas.api.types import is_bool_dtype, is_numeric_dtype
import polars as pl


ERROR_TOKENS = {"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"}


@dataclass(frozen=True)
class ResultSheetConfig:
    source_key: str
    sheet_name: str
    table_name: str
    description: str
    key_columns: str
    column_help: tuple[str, ...]
    example_lines: tuple[str, ...]
    notes: str = ""
    optional: bool = False


RESULT_SHEET_CONFIGS = (
    ResultSheetConfig(
        source_key="summary",
        sheet_name="summary",
        table_name="tbl_run_summary",
        description="Run 수준에서 각 clustering group의 핵심 결과를 한 줄씩 요약합니다.",
        key_columns="group_id, n_trials, n_clusters, selection_status",
        column_help=(
            "group_id: clustering space 또는 비교 group 식별자입니다.",
            "n_trials: 해당 그룹에 포함된 trial 수입니다.",
            "n_clusters: 최종 대표 synergy cluster 개수입니다.",
            "selection_status: K 선택이 어떤 경로로 확정되었는지 보여줍니다.",
        ),
        example_lines=(
            "예를 들어 group_id=pooled_step_nonstep 이고 selection_status=success_gap_unique 이면, pooled clustering 공간이 gap statistic이 고른 K를 그대로 최종 채택했다는 뜻입니다.",
            "group_figure_path 를 열면 해당 그룹의 대표 cluster figure 위치를 바로 찾을 수 있습니다.",
        ),
        notes="이 시트를 먼저 보고 필요한 상세 시트로 이동합니다.",
    ),
    ResultSheetConfig(
        source_key="metadata",
        sheet_name="clustering_meta",
        table_name="tbl_clustering_metadata",
        description="각 그룹의 clustering 선택 과정과 gap/중복 관련 메타데이터를 보여줍니다.",
        key_columns="group_id, selection_method, k_gap_raw, k_selected",
        column_help=(
            "selection_method: 대표 K를 선택한 기본 방법입니다.",
            "k_gap_raw: gap statistic이 처음 제안한 K입니다.",
            "k_selected: 실제로 최종 확정된 K입니다.",
            "duplicate_trial_count_by_k_json: K별 duplicate trial 수를 JSON으로 기록합니다.",
        ),
        example_lines=(
            "예를 들어 k_gap_raw=6 이고 k_selected=7 이면, 처음 제안된 K=6 대신 보정된 K=7이 최종 채택된 경우입니다.",
            "duplicate_trial_count_by_k_json 를 함께 보면 왜 K가 올라갔는지 추적할 수 있습니다.",
        ),
        notes="selection_status 와 K 관련 컬럼을 같이 읽어야 해석이 쉽습니다.",
    ),
    ResultSheetConfig(
        source_key="trial_windows",
        sheet_name="trial_windows",
        table_name="tbl_trial_window_metadata",
        description="각 trial이 어떤 analysis window로 잘렸는지와 step/nonstep 분류 결과를 보여줍니다.",
        key_columns="trial_id, analysis_step_class, analysis_window_source, analysis_window_is_surrogate",
        column_help=(
            "trial_id: subject-velocity-trial 단위 식별자입니다.",
            "analysis_step_class: step 또는 nonstep 최종 분류입니다.",
            "analysis_window_source: window 종료점이 실제 step onset인지 surrogate인지 구분합니다.",
            "analysis_window_is_surrogate: surrogate 종료점 사용 여부입니다.",
        ),
        example_lines=(
            "예를 들어 analysis_step_class=nonstep 이고 analysis_window_is_surrogate=True 이면, nonstep trial이라 실제 step onset 대신 surrogate 종료점을 사용했다는 뜻입니다.",
            "같은 subject 안에서 step/nonstep trial을 비교할 때 이 시트를 기준 truth로 삼으면 됩니다.",
        ),
        notes="paired 해석을 할 때 가장 먼저 확인해야 하는 상세 시트입니다.",
    ),
    ResultSheetConfig(
        source_key="labels",
        sheet_name="cluster_labels",
        table_name="tbl_cluster_labels",
        description="각 component가 어떤 cluster_id 로 배정되었는지 label 수준에서 보여줍니다.",
        key_columns="trial_id, component_index, cluster_id, analysis_step_class",
        column_help=(
            "component_index: trial 내부 NMF component 번호입니다.",
            "cluster_id: component가 배정된 대표 cluster 번호입니다.",
            "trial_id: 어느 trial의 component인지 식별합니다.",
            "analysis_step_class: 해당 trial의 step/nonstep class 입니다.",
        ),
        example_lines=(
            "예를 들어 같은 trial_id 안에서 component_index 가 다르지만 cluster_id 가 같으면, 서로 다른 component가 같은 대표 cluster로 묶였다는 뜻입니다.",
            "step/nonstep 차이는 analysis_step_class 컬럼으로 함께 읽으면 됩니다.",
        ),
        notes="component 단위 cluster 할당을 직접 확인할 때 사용합니다.",
    ),
    ResultSheetConfig(
        source_key="rep_W",
        sheet_name="representative_W",
        table_name="tbl_representative_w",
        description="대표 synergy W 가중치를 근육별로 long format으로 저장한 표입니다.",
        key_columns="group_id, cluster_id, muscle, W_value",
        column_help=(
            "cluster_id: 대표 synergy 번호입니다.",
            "muscle: 근육 이름입니다.",
            "W_value: 해당 근육의 상대 가중치입니다.",
            "group_id: representative W가 속한 clustering group 식별자입니다.",
        ),
        example_lines=(
            "예를 들어 cluster_id=0 에서 특정 muscle 의 W_value 가 크면, 그 대표 synergy에서 해당 근육 기여도가 상대적으로 크다고 해석할 수 있습니다.",
            "같은 cluster_id 를 group_id 별로 비교하면 clustering space 간 대표 근육 패턴 차이를 볼 수 있습니다.",
        ),
        notes="근육 기여도 해석용 대표 표입니다.",
    ),
    ResultSheetConfig(
        source_key="rep_H_long",
        sheet_name="representative_H",
        table_name="tbl_representative_h",
        description="대표 synergy H time profile을 frame 단위 long format으로 저장한 표입니다.",
        key_columns="group_id, cluster_id, frame_idx, h_value",
        column_help=(
            "cluster_id: 대표 synergy 번호입니다.",
            "frame_idx: time-normalized frame index 입니다.",
            "h_value: 해당 frame의 activation 크기입니다.",
            "group_id: representative H가 속한 clustering group 식별자입니다.",
        ),
        example_lines=(
            "예를 들어 같은 cluster_id 에서 frame_idx 가 진행되며 h_value 가 커지면, 그 구간에서 activation 이 증가한다고 볼 수 있습니다.",
            "group_id 별로 같은 cluster_id 를 비교하면 clustering space 간 시간 패턴 차이를 해석할 수 있습니다.",
        ),
        notes="대표 synergy 시간 패턴을 읽을 때 사용합니다.",
    ),
    ResultSheetConfig(
        source_key="minimal_W",
        sheet_name="minimal_W",
        table_name="tbl_minimal_units_w",
        description="trial별 최소 단위 synergy W 가중치를 근육별 long format으로 저장한 표입니다.",
        key_columns="trial_id, component_index, muscle, W_value",
        column_help=(
            "trial_id: 어느 trial의 결과인지 식별합니다.",
            "component_index: trial 내부 NMF component 번호입니다.",
            "muscle: 근육 이름입니다.",
            "W_value: 해당 근육 가중치입니다.",
        ),
        example_lines=(
            "예를 들어 같은 trial_id 안에서 component_index=0 의 W_value 패턴을 보면, 그 trial의 첫 번째 최소 단위 synergy 근육 조합을 읽을 수 있습니다.",
            "대표 cluster와 비교하려면 cluster_labels 시트의 같은 trial_id/component_index 를 함께 보면 됩니다.",
        ),
        notes="trial 단위 원본 synergy 해석용 표입니다.",
    ),
    ResultSheetConfig(
        source_key="minimal_H_long",
        sheet_name="minimal_H",
        table_name="tbl_minimal_units_h",
        description="trial별 최소 단위 synergy H time profile을 frame 단위 long format으로 저장한 표입니다.",
        key_columns="trial_id, component_index, frame_idx, h_value",
        column_help=(
            "trial_id: 어느 trial의 결과인지 식별합니다.",
            "component_index: trial 내부 NMF component 번호입니다.",
            "frame_idx: time-normalized frame index 입니다.",
            "h_value: 해당 frame activation 크기입니다.",
        ),
        example_lines=(
            "예를 들어 같은 trial_id 와 component_index 를 고정하고 frame_idx 를 따라가면, 그 component activation이 시간에 따라 어떻게 변하는지 볼 수 있습니다.",
            "trial 안 여러 component를 비교하면 activation timing 차이를 해석할 수 있습니다.",
        ),
        notes="trial 단위 activation time profile 해석용 표입니다.",
    ),
    ResultSheetConfig(
        source_key="cross_group_pairwise",
        sheet_name="cross_group_pairwise",
        table_name="tbl_cross_group_pairwise",
        description="step cluster와 nonstep cluster의 모든 representative W cosine 조합을 한 번씩 기록한 표입니다.",
        key_columns="step_cluster_id, nonstep_cluster_id, cosine_similarity, selected_in_assignment, passes_threshold, match_id",
        column_help=(
            "step_cluster_id: step 쪽 대표 cluster 번호입니다.",
            "nonstep_cluster_id: nonstep 쪽 대표 cluster 번호입니다.",
            "cosine_similarity: 두 representative W 패턴의 cosine similarity 입니다.",
            "selected_in_assignment / passes_threshold / match_id: 최종 1:1 assignment 채택 여부와 same_synergy 인정 여부를 함께 보여줍니다.",
        ),
        example_lines=(
            "예를 들어 selected_in_assignment=True 이고 passes_threshold=True 이면, 해당 step-nonstep 조합은 최종 same_synergy match로 수용된 것입니다.",
            "selected_in_assignment=True 이지만 passes_threshold=False 이면, assignment는 되었지만 최종 해석은 group_specific_synergy 입니다.",
        ),
        notes="pairwise 조합 전체를 확인할 때 쓰는 기준 표입니다.",
        optional=True,
    ),
    ResultSheetConfig(
        source_key="cross_group_matrix",
        sheet_name="cross_group_matrix",
        table_name="tbl_cross_group_matrix",
        description="행=step cluster, 열=nonstep cluster 인 cosine similarity matrix view입니다.",
        key_columns="step_cluster_id, nonstep_cluster_<id>",
        column_help=(
            "step_cluster_id: 행 기준이 되는 step cluster 번호입니다.",
            "nonstep_cluster_<id>: 각 열은 해당 nonstep cluster와의 cosine similarity를 뜻합니다.",
        ),
        example_lines=(
            "예를 들어 step_cluster_id=1 행에서 가장 큰 값을 가진 열이 그 step cluster의 최고 유사 nonstep 후보입니다.",
            "assignment 결과와 다를 수 있으므로 최종 해석은 cross_group_decision 시트를 함께 확인합니다.",
        ),
        notes="전체 유사도 구조를 한눈에 볼 때 사용합니다.",
        optional=True,
    ),
    ResultSheetConfig(
        source_key="cross_group_decision",
        sheet_name="cross_group_decision",
        table_name="tbl_cross_group_decision",
        description="cluster당 정확히 1행으로 최종 same_synergy / group_specific_synergy 판정을 기록한 표입니다.",
        key_columns="group_id, cluster_id, final_label, match_id, assigned_partner_cluster_id, assigned_cosine_similarity, best_partner_cluster_id, best_partner_cosine_similarity",
        column_help=(
            "final_label: 최종 해석 라벨이며 same_synergy 또는 group_specific_synergy 만 허용됩니다.",
            "assigned_partner_cluster_id / assigned_cosine_similarity: Hungarian assignment 결과를 그대로 보존합니다.",
            "best_partner_cluster_id / best_partner_cosine_similarity: threshold 미통과 또는 미배정 cluster도 최대 cosine 정보를 잃지 않게 남깁니다.",
        ),
        example_lines=(
            "예를 들어 final_label=same_synergy 이고 match_id가 채워져 있으면, 그 cluster는 반대 그룹 cluster와 같은 synergy family로 해석합니다.",
            "final_label=group_specific_synergy 이더라도 assigned_cosine_similarity 또는 best_partner_cosine_similarity 를 보고 얼마나 가까웠는지 확인할 수 있습니다.",
        ),
        notes="최종 biological interpretation은 이 시트를 기준으로 읽습니다.",
        optional=True,
    ),
    ResultSheetConfig(
        source_key="cross_group_summary",
        sheet_name="cross_group_summary",
        table_name="tbl_cross_group_summary",
        description="cross-group representative W 비교 결과를 한 줄 요약으로 보여주는 표입니다.",
        key_columns="step_cluster_count, nonstep_cluster_count, accepted_same_synergy_match_count, group_specific_step_cluster_count, group_specific_nonstep_cluster_count, threshold",
        column_help=(
            "step_cluster_count / nonstep_cluster_count: 비교에 참여한 cluster 수입니다.",
            "accepted_same_synergy_match_count: threshold를 통과해 수용된 same_synergy match 개수입니다.",
            "group_specific_step_cluster_count / group_specific_nonstep_cluster_count: 최종 group_specific_synergy 로 남은 cluster 수입니다.",
            "threshold: same_synergy 인정에 사용한 cosine 기준값입니다.",
        ),
        example_lines=(
            "예를 들어 accepted_same_synergy_match_count=2 이면, step-nonstep 사이에서 2개의 대표 synergy family가 공유되었다는 뜻입니다.",
            "group_specific_step_cluster_count 가 크면, step 전략에만 남는 representative W 패턴이 상대적으로 많다고 해석할 수 있습니다.",
        ),
        notes="먼저 전체 개요를 보고 세부 시트로 내려갈 때 사용합니다.",
        optional=True,
    ),
)


GUIDE_SHEET_CONFIG = ResultSheetConfig(
    source_key="table_guide",
    sheet_name="table_guide",
    table_name="tbl_table_guide",
    description="이 워크북의 모든 Excel Table 위치와 핵심 읽는 법을 한눈에 정리합니다.",
    key_columns="table_name, sheet_name, key_columns",
    column_help=(
        "table_name: 실제 Excel Table 이름입니다.",
        "sheet_name: 표가 들어 있는 시트 이름입니다.",
        "key_columns: 먼저 확인해야 하는 핵심 컬럼입니다.",
        "column_guide: 핵심 컬럼이 각각 무엇을 뜻하는지 요약합니다.",
    ),
    example_lines=(
        "예를 들어 trial_windows 시트가 궁금하면 sheet_name=trial_windows 행을 찾고, key_columns 와 column_guide 를 먼저 읽으면 됩니다.",
        "그 다음 table_range 를 따라가면 실제 표 위치를 바로 찾을 수 있습니다.",
    ),
    notes="워크북 입문용 인덱스 시트입니다.",
)


@dataclass(frozen=True)
class TablePlacement:
    table_name: str
    sheet_name: str
    start_row: int
    start_col: int
    frame: pd.DataFrame
    description: str
    key_columns: str
    column_guide: str
    notes: str = ""


def _pl_to_pandas(frame: pl.DataFrame) -> pd.DataFrame:
    return pd.DataFrame(frame.to_dicts(), columns=frame.columns)


def _sheet_lines(config: ResultSheetConfig) -> list[str]:
    lines = [f"[목적] {config.description}", "[핵심 컬럼]"]
    lines.extend(f"{index}. {line}" for index, line in enumerate(config.column_help, start=1))
    lines.append("[예시]")
    lines.extend(config.example_lines)
    if config.notes:
        lines.append("[읽는 팁]")
        lines.append(config.notes)
    return lines


def _write_text_block(sheet, lines: list[str], start_row: int = 1, start_col: int = 1) -> int:
    row = start_row
    for line in lines:
        cell = sheet.cell(row=row, column=start_col, value=line)
        if line.startswith("["):
            cell.font = Font(bold=True)
        row += 1
    return row


def _coerce_cell_value(value: Any, *, force_text: bool) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if force_text:
        return str(value)
    return value


def _table_ref(start_row: int, start_col: int, frame: pd.DataFrame) -> str:
    end_row = start_row + len(frame.index)
    end_col = start_col + len(frame.columns) - 1
    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"


def _infer_text_columns(frame: pd.DataFrame) -> set[str]:
    text_columns: set[str] = set()
    for column_name in frame.columns:
        series = frame[column_name]
        if is_bool_dtype(series) or not is_numeric_dtype(series):
            text_columns.add(column_name)
    return text_columns


def _write_table(sheet, placement: TablePlacement) -> tuple[str, int]:
    frame = placement.frame
    text_columns = _infer_text_columns(frame)
    for offset, column_name in enumerate(frame.columns, start=placement.start_col):
        header_cell = sheet.cell(row=placement.start_row, column=offset, value=str(column_name))
        header_cell.font = Font(bold=True)
    for row_offset, row in enumerate(frame.itertuples(index=False), start=1):
        for col_offset, (column_name, value) in enumerate(zip(frame.columns, row, strict=True), start=placement.start_col):
            force_text = column_name in text_columns
            cell = sheet.cell(
                row=placement.start_row + row_offset,
                column=col_offset,
                value=_coerce_cell_value(value, force_text=force_text),
            )
            if force_text:
                cell.number_format = "@"
    ref = _table_ref(placement.start_row, placement.start_col, frame)
    table = Table(displayName=placement.table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    for index, column_name in enumerate(frame.columns, start=placement.start_col):
        width = max(len(str(column_name)), 14)
        for value in frame.iloc[:, index - placement.start_col].astype(str).tolist():
            width = min(max(width, len(value) + 2), 60)
        sheet.column_dimensions[get_column_letter(index)].width = width
    return ref, placement.start_row + len(frame.index) + 3


def _placeholder_frame(frame: pd.DataFrame, message: str) -> pd.DataFrame:
    if not frame.empty:
        return frame
    if list(frame.columns):
        row = {column: "" for column in frame.columns}
        row[frame.columns[0]] = message
        return pd.DataFrame([row], columns=frame.columns)
    return pd.DataFrame([{"message": message}])


def _table_guide_frame(table_rows: list[dict[str, str]]) -> pd.DataFrame:
    return _pl_to_pandas(
        pl.DataFrame(
            table_rows,
            schema=[
                "table_name",
                "sheet_name",
                "table_range",
                "description",
                "key_columns",
                "column_guide",
                "notes",
            ],
            orient="row",
        )
    )


def _sheet_frames(summary_frame: pd.DataFrame, aggregate_frames: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    return {"summary": summary_frame, **aggregate_frames}


def _active_result_sheet_configs(frames: dict[str, pd.DataFrame]) -> tuple[ResultSheetConfig, ...]:
    return tuple(
        config
        for config in RESULT_SHEET_CONFIGS
        if not config.optional or config.source_key in frames
    )


def write_results_interpretation_workbook(
    path: Path,
    summary_frame: pd.DataFrame,
    aggregate_frames: dict[str, pd.DataFrame],
) -> Path:
    """Write a run-level interpretation workbook and validate it after reopening."""
    frames = _sheet_frames(summary_frame, aggregate_frames)
    active_configs = _active_result_sheet_configs(frames)
    workbook = Workbook()
    table_info: list[dict[str, str]] = []
    guide_sheet = None

    for index, config in enumerate(active_configs):
        raw_frame = frames.get(config.source_key, pd.DataFrame())
        frame = _placeholder_frame(raw_frame, f"No rows were exported for {config.sheet_name}.")
        sheet = workbook.active if index == 0 else workbook.create_sheet(config.sheet_name)
        sheet.title = config.sheet_name
        start_row = _write_text_block(sheet, _sheet_lines(config), start_row=1, start_col=1) + 1
        placement = TablePlacement(
            table_name=config.table_name,
            sheet_name=config.sheet_name,
            start_row=start_row,
            start_col=1,
            frame=frame,
            description=config.description,
            key_columns=config.key_columns,
            column_guide=" | ".join(config.column_help),
            notes=config.notes,
        )
        ref, _ = _write_table(sheet, placement)
        table_info.append(
            {
                "table_name": config.table_name,
                "sheet_name": config.sheet_name,
                "table_range": ref,
                "description": config.description,
                "key_columns": config.key_columns,
                "column_guide": " | ".join(config.column_help),
                "notes": config.notes,
            }
        )

    guide_sheet = workbook.create_sheet(GUIDE_SHEET_CONFIG.sheet_name)
    guide_start_row = _write_text_block(guide_sheet, _sheet_lines(GUIDE_SHEET_CONFIG), start_row=1, start_col=1) + 1
    guide_row = {
        "table_name": GUIDE_SHEET_CONFIG.table_name,
        "sheet_name": GUIDE_SHEET_CONFIG.sheet_name,
        "table_range": "pending",
        "description": GUIDE_SHEET_CONFIG.description,
        "key_columns": GUIDE_SHEET_CONFIG.key_columns,
        "column_guide": " | ".join(GUIDE_SHEET_CONFIG.column_help),
        "notes": GUIDE_SHEET_CONFIG.notes,
    }
    guide_frame = _table_guide_frame(table_info + [guide_row])
    guide_placement = TablePlacement(
        table_name=GUIDE_SHEET_CONFIG.table_name,
        sheet_name=GUIDE_SHEET_CONFIG.sheet_name,
        start_row=guide_start_row,
        start_col=1,
        frame=guide_frame,
        description=GUIDE_SHEET_CONFIG.description,
        key_columns=GUIDE_SHEET_CONFIG.key_columns,
        column_guide=" | ".join(GUIDE_SHEET_CONFIG.column_help),
        notes=GUIDE_SHEET_CONFIG.notes,
    )
    guide_ref, _ = _write_table(guide_sheet, guide_placement)
    guide_range_col = guide_frame.columns.get_loc("table_range") + 1
    guide_self_row = guide_placement.start_row + len(guide_frame.index)
    guide_sheet.cell(row=guide_self_row, column=guide_range_col, value=guide_ref).number_format = "@"

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    validate_results_interpretation_workbook(path)
    return path


def validate_results_interpretation_workbook(path: Path) -> dict[str, Any]:
    """Reopen the workbook and confirm every interpretation sheet is present and documented."""
    required_configs = tuple(config for config in RESULT_SHEET_CONFIGS if not config.optional)
    expected_sheets = {config.sheet_name for config in required_configs} | {GUIDE_SHEET_CONFIG.sheet_name}
    expected_tables = {config.sheet_name: {config.table_name} for config in required_configs}
    expected_tables[GUIDE_SHEET_CONFIG.sheet_name] = {GUIDE_SHEET_CONFIG.table_name}
    issues = {"errors": [], "blanks": [], "tables": []}
    workbook = load_workbook(path)
    try:
        sheet_names = set(workbook.sheetnames)
        for sheet_name in sorted(expected_sheets - sheet_names):
            issues["tables"].append((sheet_name, "missing_sheet"))

        actual_workbook_tables: set[tuple[str, str]] = set()
        for sheet_name, table_names in expected_tables.items():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            actual_names = set(sheet.tables.keys())
            actual_workbook_tables.update((sheet_name, table_name) for table_name in actual_names)
            for table_name in sorted(table_names - actual_names):
                issues["tables"].append((table_name, "missing_table"))

            for table_name in sorted(actual_names):
                min_col, min_row, max_col, max_row = range_boundaries(sheet.tables[table_name].ref)
                for row in sheet.iter_rows(
                    min_row=min_row + 1,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                ):
                    for value in row:
                        if isinstance(value, str) and value in ERROR_TOKENS:
                            issues["errors"].append((sheet_name, table_name, value))

            guide_values = [
                sheet.cell(row=row_idx, column=1).value
                for row_idx in range(1, min(sheet.max_row, 20) + 1)
            ]
            for label in ("[목적]", "[핵심 컬럼]", "[예시]"):
                if not any(isinstance(value, str) and value.startswith(label) for value in guide_values):
                    issues["blanks"].append((sheet_name, label))

        guide_sheet = workbook[GUIDE_SHEET_CONFIG.sheet_name] if GUIDE_SHEET_CONFIG.sheet_name in workbook.sheetnames else None
        guide_table = None if guide_sheet is None else guide_sheet.tables.get(GUIDE_SHEET_CONFIG.table_name)
        if guide_table is None:
            pass
        else:
            min_col, min_row, max_col, max_row = range_boundaries(guide_table.ref)
            rows = list(
                guide_sheet.iter_rows(
                    min_row=min_row + 1,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
            )
            documented_tables: set[tuple[str, str]] = set()
            for row in rows:
                if row[0] in (None, ""):
                    continue
                documented_tables.add((str(row[1]), str(row[0])))
                if any(cell in ERROR_TOKENS for cell in row if isinstance(cell, str)):
                    issues["errors"].append((GUIDE_SHEET_CONFIG.sheet_name, row[0], "error_token"))
                if row[3] in (None, "") or row[5] in (None, ""):
                    issues["blanks"].append((GUIDE_SHEET_CONFIG.sheet_name, row[0]))
            for table_entry in sorted(actual_workbook_tables - documented_tables):
                issues["tables"].append((table_entry[1], "missing_guide_row"))
    finally:
        workbook.close()

    if any(issues.values()):
        raise ValueError(f"Interpretation workbook validation failed for {path}: {issues}")
    return {
        "workbook_path": str(path),
        "sheet_count": len(expected_sheets),
        "table_count": sum(len(table_names) for table_names in expected_tables.values()),
        "engine": "openpyxl",
        "excel_ui_visual_qa": "skipped_desktop_excel_unavailable",
        "fallback_reason": "openpyxl_default_wsl_headless_environment",
        "issues": issues,
    }
