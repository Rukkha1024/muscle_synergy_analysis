"""Contract tests for the clustering audit workbook export."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import numpy as np
import pytest

from src.synergy_stats.artifacts import export_results
from src.synergy_stats.clustering import SubjectFeatureResult
from src.synergy_stats.excel_audit import (
    build_audit_tables,
    validate_clustering_audit_workbook,
    write_clustering_audit_workbook,
)
from src.synergy_stats.excel_results import validate_results_interpretation_workbook


def _mock_cluster_group_results() -> dict[str, dict]:
    duplicate_row = {
        "subject": "S01",
        "velocity": 1,
        "trial_num": 2,
        "trial_id": "S01_v1_T2",
        "trial_key": ("S01", 1, 2),
        "n_synergies_in_trial": 4,
        "duplicate_cluster_labels": [1, 3],
        "duplicate_component_indexes": [0, 1, 2, 3],
        "duplicate_cluster_count": 2,
        "duplicate_component_count": 4,
        "duplicate_cluster_details": [
            {"cluster_id": 1, "component_indexes": [0, 2], "component_count": 2},
            {"cluster_id": 3, "component_indexes": [1, 3], "component_count": 2},
        ],
    }
    return {
        "global_step": {
            "cluster_result": {
                "selection_status": "success_gap_unique",
                "k_gap_raw": 10,
                "k_selected": 10,
                "k_min_unique": 10,
                "gap_by_k": {9: 1.0, 10: 1.2},
                "gap_sd_by_k": {9: 0.1, 10: 0.1},
                "observed_objective_by_k": {9: 90.0, 10: 80.0},
                "feasible_objective_by_k": {9: 91.0, 10: 80.0},
                "duplicate_trial_count_by_k": {9: 2, 10: 0},
                "duplicate_trial_evidence_by_k": {9: [duplicate_row], 10: []},
            }
        },
        "global_nonstep": {
            "cluster_result": {
                "selection_status": "success_gap_escalated_unique",
                "k_gap_raw": 6,
                "k_selected": 7,
                "k_min_unique": 7,
                "gap_by_k": {6: 1.1, 7: 1.0},
                "gap_sd_by_k": {6: 0.1, 7: 0.1},
                "observed_objective_by_k": {6: 60.0, 7: 55.0},
                "feasible_objective_by_k": {6: "", 7: 55.0},
                "duplicate_trial_count_by_k": {6: 1, 7: 0},
                "duplicate_trial_evidence_by_k": {6: [duplicate_row], 7: []},
            }
        },
    }


def test_build_audit_tables_returns_selection_and_duplicate_frames() -> None:
    """Audit tables should flatten selection, K diagnostics, and duplicate details."""
    frames = build_audit_tables(_mock_cluster_group_results())

    assert list(frames["selection_summary"]["group_id"]) == ["global_step", "global_nonstep"]
    assert "summary_text" in frames["selection_summary"].columns
    assert len(frames["k_audit"]) == 4
    assert len(frames["duplicate_trial_summary"]) == 2
    assert len(frames["duplicate_cluster_detail"]) == 4


def test_write_clustering_audit_workbook_creates_expected_sheets_and_tables(tmp_path: Path) -> None:
    """Workbook export should include guide text, expected sheets, and Excel tables."""
    output_path = tmp_path / "clustering_audit.xlsx"
    write_clustering_audit_workbook(output_path, _mock_cluster_group_results())

    workbook = load_workbook(output_path)
    try:
        assert workbook.sheetnames == ["summary", "duplicates", "table_guide"]
        assert workbook["summary"]["A2"].value == "[표 읽는 순서]"
        assert workbook["summary"]["A8"].value == "[예시]"
        assert workbook["summary"]["A11"].value == "[실행 환경 메모]"
        assert workbook["summary"]["A12"].value == "Workbook 엔진: openpyxl 기본값."
        assert workbook["duplicates"]["A2"].value == "[표 읽는 순서]"
        assert workbook["duplicates"]["A7"].value == "[예시]"
        assert "trial_id=S01_v1_T2" in workbook["duplicates"]["A8"].value
        assert workbook["table_guide"]["A2"].value == "[표 읽는 순서]"
        assert workbook["table_guide"]["A8"].value == "[예시]"
        assert "tbl_duplicate_trial_summary" in workbook["table_guide"]["A9"].value
        assert set(workbook["summary"].tables.keys()) == {
            "tbl_clustering_selection_summary",
            "tbl_clustering_k_audit",
        }
        assert set(workbook["duplicates"].tables.keys()) == {
            "tbl_duplicate_trial_summary",
            "tbl_duplicate_cluster_detail",
        }
        assert set(workbook["table_guide"].tables.keys()) == {"tbl_table_guide"}
        guide_table = workbook["table_guide"].tables["tbl_table_guide"]
        guide_min_col, guide_min_row, _, _ = range_boundaries(guide_table.ref)
        assert workbook["table_guide"].cell(row=guide_min_row + 5, column=guide_min_col).value == "tbl_table_guide"
    finally:
        workbook.close()


def test_write_clustering_audit_workbook_includes_placeholder_duplicate_rows_when_none_exist(tmp_path: Path) -> None:
    """Workbook should keep duplicate tables readable even when no duplicate rows exist."""
    cluster_group_results = _mock_cluster_group_results()
    for payload in cluster_group_results.values():
        payload["cluster_result"]["duplicate_trial_evidence_by_k"] = {10: [], 9: []} if payload["cluster_result"]["k_gap_raw"] == 10 else {6: [], 7: []}
        payload["cluster_result"]["duplicate_trial_count_by_k"] = {10: 0, 9: 0} if payload["cluster_result"]["k_gap_raw"] == 10 else {6: 0, 7: 0}

    output_path = tmp_path / "clustering_audit_empty_duplicates.xlsx"
    write_clustering_audit_workbook(output_path, cluster_group_results)

    workbook = load_workbook(output_path)
    try:
        duplicates_sheet = workbook["duplicates"]
        trial_table = duplicates_sheet.tables["tbl_duplicate_trial_summary"]
        cluster_table = duplicates_sheet.tables["tbl_duplicate_cluster_detail"]
        trial_min_col, trial_min_row, _, _ = range_boundaries(trial_table.ref)
        cluster_min_col, cluster_min_row, _, _ = range_boundaries(cluster_table.ref)
        assert duplicates_sheet.cell(row=trial_min_row + 1, column=trial_min_col + 5).value == (
            "No duplicate trials were found across the audited K range."
        )
        assert duplicates_sheet.cell(row=cluster_min_row + 1, column=cluster_min_col + 5).value == (
            "No duplicate cluster details were found across the audited K range."
        )
    finally:
        workbook.close()


def test_validate_clustering_audit_workbook_reports_missing_sheet_without_keyerror(tmp_path: Path) -> None:
    """Validation should report a missing sheet as workbook issues instead of raising KeyError."""
    output_path = tmp_path / "clustering_audit_missing_duplicates.xlsx"
    write_clustering_audit_workbook(output_path, _mock_cluster_group_results())

    workbook = load_workbook(output_path)
    try:
        del workbook["duplicates"]
        workbook.save(output_path)
    finally:
        workbook.close()

    with pytest.raises(ValueError, match="missing_sheet"):
        validate_clustering_audit_workbook(output_path)


def test_validate_clustering_audit_workbook_reports_missing_table_guide_without_keyerror(tmp_path: Path) -> None:
    """Validation should report a missing table_guide sheet as workbook issues instead of raising KeyError."""
    output_path = tmp_path / "clustering_audit_missing_table_guide.xlsx"
    write_clustering_audit_workbook(output_path, _mock_cluster_group_results())

    workbook = load_workbook(output_path)
    try:
        del workbook["table_guide"]
        workbook.save(output_path)
    finally:
        workbook.close()

    with pytest.raises(ValueError, match="missing_sheet"):
        validate_clustering_audit_workbook(output_path)


def test_export_results_records_clustering_audit_workbook_path(tmp_path: Path, monkeypatch) -> None:
    """Artifact export should write the workbook and expose its path in context artifacts."""
    import src.synergy_stats.artifacts as artifacts_module

    feature_rows = []
    w_muscle = np.array([[1.0, 0.0], [0.0, 1.0], [0.5, 0.5]], dtype=np.float32)
    h_time = np.tile(np.array([[1.0, 0.5]], dtype=np.float32), (6, 1))
    for group_id, subject in (("global_step", "S01"), ("global_nonstep", "S02")):
        step_class = "step" if group_id == "global_step" else "nonstep"
        feature_rows.append(
            SubjectFeatureResult(
                subject=subject,
                velocity=1,
                trial_num=1,
                bundle=SimpleNamespace(
                    W_muscle=w_muscle,
                    H_time=h_time,
                    meta={
                        "analysis_selected_group": True,
                        "analysis_step_class": step_class,
                        "analysis_is_step": group_id == "global_step",
                        "analysis_is_nonstep": group_id == "global_nonstep",
                    },
                ),
            )
        )

    cluster_result = {
        "status": "success",
        "n_clusters": 2,
        "labels": np.array([0, 1], dtype=np.int32),
        "inertia": 0.1,
        "duplicate_trials": [],
        "algorithm_used": "mock_kmeans",
        "selection_method": "gap_statistic",
        "selection_status": "success_gap_unique",
        "duplicate_resolution": "none",
        "require_zero_duplicate_solution": True,
        "k_lb": 2,
        "k_gap_raw": 2,
        "k_selected": 2,
        "k_min_unique": 2,
        "repeats": 1,
        "gap_ref_n": 2,
        "gap_ref_restarts": 1,
        "uniqueness_candidate_restarts": 1,
        "gap_by_k": {2: 1.0},
        "gap_sd_by_k": {2: 0.1},
        "observed_objective_by_k": {2: 0.1},
        "feasible_objective_by_k": {2: 0.1},
        "duplicate_trial_count_by_k": {2: 0},
        "duplicate_trial_evidence_by_k": {2: []},
        "sample_map": [
            {
                "group_id": "global_step",
                "subject": "S01",
                "velocity": 1,
                "trial_num": 1,
                "component_index": 0,
                "trial_key": ("S01", 1, 1),
                "trial_id": "S01_v1_T1",
            },
            {
                "group_id": "global_step",
                "subject": "S01",
                "velocity": 1,
                "trial_num": 1,
                "component_index": 1,
                "trial_key": ("S01", 1, 1),
                "trial_id": "S01_v1_T1",
            },
        ],
    }

    def _fake_group_figure(*args, **kwargs):
        output_path = kwargs["output_path"]
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b"figure")

    def _fake_trial_figure(*args, **kwargs):
        output_path = kwargs["output_path"]
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b"trial-figure")

    monkeypatch.setattr(artifacts_module, "save_group_cluster_figure", _fake_group_figure)
    monkeypatch.setattr(artifacts_module, "save_trial_nmf_figure", _fake_trial_figure)

    context = {
        "config": {
            "runtime": {
                "output_dir": str(tmp_path / "run_output"),
                "final_parquet_path": str(tmp_path / "final.parquet"),
            },
            "muscles": {"names": ["M1", "M2", "M3"]},
            "synergy_clustering": {
                "representative": {"h_output_interpolation": {"target_windows": 10}}
            },
            "figures": {"format": "png", "dpi": 150, "overview_columns": 2},
        },
        "feature_rows": feature_rows,
        "cluster_group_results": {
            "global_step": {
                "group_id": "global_step",
                "feature_rows": [feature_rows[0]],
                "cluster_result": cluster_result,
            },
            "global_nonstep": {
                "group_id": "global_nonstep",
                "feature_rows": [feature_rows[1]],
                "cluster_result": {
                    **cluster_result,
                    "sample_map": [
                        {
                            "group_id": "global_nonstep",
                            "subject": "S02",
                            "velocity": 1,
                            "trial_num": 1,
                            "component_index": 0,
                            "trial_key": ("S02", 1, 1),
                            "trial_id": "S02_v1_T1",
                        },
                        {
                            "group_id": "global_nonstep",
                            "subject": "S02",
                            "velocity": 1,
                            "trial_num": 1,
                            "component_index": 1,
                            "trial_key": ("S02", 1, 1),
                            "trial_id": "S02_v1_T1",
                        },
                    ],
                },
            },
        },
        "artifacts": {"steps": []},
    }

    updated = export_results(context)

    workbook_path = Path(updated["artifacts"]["clustering_audit_workbook_path"])
    interpretation_path = Path(updated["artifacts"]["results_interpretation_workbook_path"])
    assert workbook_path.exists()
    assert interpretation_path.exists()
    assert updated["artifacts"]["clustering_audit_workbook_validation"]["engine"] == "openpyxl"
    assert updated["artifacts"]["results_interpretation_workbook_validation"]["engine"] == "openpyxl"
    workbook = load_workbook(workbook_path)
    try:
        assert workbook.sheetnames == ["summary", "duplicates", "table_guide"]
    finally:
        workbook.close()
    interpretation_book = load_workbook(interpretation_path)
    try:
        assert interpretation_book.sheetnames == [
            "summary",
            "clustering_meta",
            "trial_windows",
            "cluster_labels",
            "representative_W",
            "representative_H",
            "minimal_W",
            "minimal_H",
            "table_guide",
        ]
        assert interpretation_book["trial_windows"]["A2"].value == "[핵심 컬럼]"
        assert interpretation_book["trial_windows"]["A7"].value == "[예시]"
        assert interpretation_book["table_guide"]["A2"].value == "[핵심 컬럼]"
        assert set(interpretation_book["table_guide"].tables.keys()) == {"tbl_table_guide"}
    finally:
        interpretation_book.close()


def test_validate_results_interpretation_workbook_catches_error_tokens(tmp_path: Path, monkeypatch) -> None:
    """Interpretation workbook validation should fail when a table cell contains an Excel error token."""
    import src.synergy_stats.artifacts as artifacts_module

    feature_rows = []
    w_muscle = np.array([[1.0, 0.0], [0.0, 1.0], [0.5, 0.5]], dtype=np.float32)
    h_time = np.tile(np.array([[1.0, 0.5]], dtype=np.float32), (6, 1))
    for group_id, subject in (("global_step", "S01"), ("global_nonstep", "S02")):
        step_class = "step" if group_id == "global_step" else "nonstep"
        feature_rows.append(
            SubjectFeatureResult(
                subject=subject,
                velocity=1,
                trial_num=1,
                bundle=SimpleNamespace(
                    W_muscle=w_muscle,
                    H_time=h_time,
                    meta={
                        "analysis_selected_group": True,
                        "analysis_step_class": step_class,
                        "analysis_is_step": group_id == "global_step",
                        "analysis_is_nonstep": group_id == "global_nonstep",
                    },
                ),
            )
        )

    cluster_result = {
        "status": "success",
        "n_clusters": 2,
        "labels": np.array([0, 1], dtype=np.int32),
        "inertia": 0.1,
        "duplicate_trials": [],
        "algorithm_used": "mock_kmeans",
        "selection_method": "gap_statistic",
        "selection_status": "success_gap_unique",
        "duplicate_resolution": "none",
        "require_zero_duplicate_solution": True,
        "k_lb": 2,
        "k_gap_raw": 2,
        "k_selected": 2,
        "k_min_unique": 2,
        "repeats": 1,
        "gap_ref_n": 2,
        "gap_ref_restarts": 1,
        "uniqueness_candidate_restarts": 1,
        "gap_by_k": {2: 1.0},
        "gap_sd_by_k": {2: 0.1},
        "observed_objective_by_k": {2: 0.1},
        "feasible_objective_by_k": {2: 0.1},
        "duplicate_trial_count_by_k": {2: 0},
        "duplicate_trial_evidence_by_k": {2: []},
        "sample_map": [
            {
                "group_id": "global_step",
                "subject": "S01",
                "velocity": 1,
                "trial_num": 1,
                "component_index": 0,
                "trial_key": ("S01", 1, 1),
                "trial_id": "S01_v1_T1",
            },
            {
                "group_id": "global_step",
                "subject": "S01",
                "velocity": 1,
                "trial_num": 1,
                "component_index": 1,
                "trial_key": ("S01", 1, 1),
                "trial_id": "S01_v1_T1",
            },
        ],
    }

    def _fake_group_figure(*args, **kwargs):
        output_path = kwargs["output_path"]
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b"figure")

    def _fake_trial_figure(*args, **kwargs):
        output_path = kwargs["output_path"]
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b"trial-figure")

    monkeypatch.setattr(artifacts_module, "save_group_cluster_figure", _fake_group_figure)
    monkeypatch.setattr(artifacts_module, "save_trial_nmf_figure", _fake_trial_figure)

    context = {
        "config": {
            "runtime": {
                "output_dir": str(tmp_path / "run_output"),
                "final_parquet_path": str(tmp_path / "final.parquet"),
            },
            "muscles": {"names": ["M1", "M2", "M3"]},
            "synergy_clustering": {
                "representative": {"h_output_interpolation": {"target_windows": 10}}
            },
            "figures": {"format": "png", "dpi": 150, "overview_columns": 2},
        },
        "feature_rows": feature_rows,
        "cluster_group_results": {
            "global_step": {
                "group_id": "global_step",
                "feature_rows": [feature_rows[0]],
                "cluster_result": cluster_result,
            },
            "global_nonstep": {
                "group_id": "global_nonstep",
                "feature_rows": [feature_rows[1]],
                "cluster_result": {
                    **cluster_result,
                    "sample_map": [
                        {
                            "group_id": "global_nonstep",
                            "subject": "S02",
                            "velocity": 1,
                            "trial_num": 1,
                            "component_index": 0,
                            "trial_key": ("S02", 1, 1),
                            "trial_id": "S02_v1_T1",
                        },
                        {
                            "group_id": "global_nonstep",
                            "subject": "S02",
                            "velocity": 1,
                            "trial_num": 1,
                            "component_index": 1,
                            "trial_key": ("S02", 1, 1),
                            "trial_id": "S02_v1_T1",
                        },
                    ],
                },
            },
        },
        "artifacts": {"steps": []},
    }

    updated = export_results(context)
    interpretation_path = Path(updated["artifacts"]["results_interpretation_workbook_path"])
    workbook = load_workbook(interpretation_path)
    try:
        trial_windows_sheet = workbook["trial_windows"]
        trial_table = trial_windows_sheet.tables["tbl_trial_window_metadata"]
        trial_min_col, trial_min_row, _, _ = range_boundaries(trial_table.ref)
        trial_windows_sheet.cell(row=trial_min_row + 1, column=trial_min_col).value = "#REF!"
        workbook.save(interpretation_path)
    finally:
        workbook.close()

    with pytest.raises(ValueError, match="#REF!"):
        validate_results_interpretation_workbook(interpretation_path)
