"""End-to-end contract tests for single-parquet, mode-only outputs."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
import pytest
import yaml

from src.synergy_stats.single_parquet import load_single_parquet_bundle
from tests.helpers import read_final_parquet, repo_python


def _write_mode_test_configs(
    fixture_bundle: dict[str, object],
    tmp_path: Path,
    *,
    mode: str,
) -> tuple[Path, Path]:
    synergy_cfg_path = tmp_path / f"synergy_stats_{mode}.yaml"
    synergy_cfg_path.write_text(
        "\n".join(
            [
                "feature_extractor:",
                "  type: nmf",
                "  nmf:",
                "    backend: sklearn_nmf",
                "    vaf_threshold: 0.9",
                "    max_components_to_try: 4",
                "    fit_params:",
                "      max_iter: 1000",
                "      tol: 0.0001",
                "synergy_analysis:",
                f"  mode: {mode}",
                "synergy_clustering:",
                "  algorithm: sklearn_kmeans",
                "  selection_method: gap_statistic",
                "  require_zero_duplicate_solution: true",
                "  duplicate_resolution: none",
                "  max_clusters: 4",
                "  max_iter: 100",
                "  repeats: 10",
                "  gap_ref_n: 5",
                "  gap_ref_restarts: 3",
                "  uniqueness_candidate_restarts: 10",
                "  random_state: 7",
                "  representative:",
                "    h_output_interpolation:",
                "      target_windows: 100",
                "figures:",
                "  format: png",
                "  dpi: 120",
                "  overview_columns: 2",
            ]
        )
        + "\n",
        encoding="utf-8-sig",
    )

    fixture_global_cfg = yaml.safe_load(
        Path(fixture_bundle["global_config"]).read_text(encoding="utf-8-sig")
    )
    fixture_global_cfg["configs"]["synergy_stats"] = str(synergy_cfg_path)
    global_cfg_path = tmp_path / f"global_config_{mode}.yaml"
    global_cfg_path.write_text(
        yaml.safe_dump(fixture_global_cfg, sort_keys=False, allow_unicode=True),
        encoding="utf-8-sig",
    )
    return synergy_cfg_path, global_cfg_path


def _run_fixture_mode(
    repo_root: Path,
    fixture_bundle: dict[str, object],
    tmp_path: Path,
    *,
    mode: str,
    run_name: str,
):
    _, global_cfg_path = _write_mode_test_configs(fixture_bundle, tmp_path, mode=mode)
    run_dir = tmp_path / run_name
    result = repo_python(
        repo_root,
        "main.py",
        "--config",
        str(global_cfg_path),
        "--mode",
        mode,
        "--out",
        str(run_dir),
        "--overwrite",
        timeout=180,
    )
    return run_dir, result


def test_both_mode_writes_single_parquet_aliases_and_mode_only_outputs(
    repo_root: Path,
    fixture_bundle: dict[str, object],
    tmp_path: Path,
) -> None:
    """A both-mode run should write alias parquet sources and mode-specific outputs only."""
    main_path = repo_root / "main.py"
    if not main_path.exists():
        pytest.xfail("main.py is not implemented yet; end-to-end contract is staged.")

    run_dir, result = _run_fixture_mode(
        repo_root,
        fixture_bundle,
        tmp_path,
        mode="both",
        run_name="fixture_run",
    )
    if result.returncode != 0:
        pytest.fail(
            "Fixture run failed.\n"
            f"stdout:\n{result.stdout}\n"
            f"stderr:\n{result.stderr}"
        )

    final_parquet_alias = repo_root / "outputs" / "final.parquet"
    final_trialwise_alias = repo_root / "outputs" / "final_trialwise.parquet"
    final_concatenated_alias = repo_root / "outputs" / "final_concatenated.parquet"
    manifest_path = run_dir / "run_manifest.json"
    methods_manifest_path = run_dir / "analysis_methods_manifest.json"

    assert manifest_path.exists()
    assert methods_manifest_path.exists()
    assert final_parquet_alias.exists()
    assert final_trialwise_alias.exists()
    assert final_concatenated_alias.exists()
    assert not (run_dir / "parquet").exists()
    assert not (run_dir / "final.parquet").exists()
    assert not (run_dir / "clustering_audit.xlsx").exists()
    assert not (run_dir / "results_interpretation.xlsx").exists()
    assert not list(run_dir.rglob("*.csv"))

    with manifest_path.open("r", encoding="utf-8-sig") as handle:
        manifest = json.load(handle)
    assert manifest["selected_mode"] == "both"
    assert manifest["analysis_modes"] == ["trialwise", "concatenated"]
    assert Path(manifest["combined_final_parquet_path"]) == final_parquet_alias

    with methods_manifest_path.open("r", encoding="utf-8-sig") as handle:
        methods_manifest = json.load(handle)
    assert methods_manifest["analysis_modes"] == ["trialwise", "concatenated"]
    assert set(methods_manifest["modes"]) == {"trialwise", "concatenated"}

    combined_df = read_final_parquet(final_parquet_alias)
    assert "artifact_kind" in combined_df.columns
    assert set(combined_df["artifact_kind"].unique().to_list()) >= {
        "final_summary",
        "metadata",
        "labels",
        "rep_W",
        "rep_H_long",
        "minimal_W",
        "minimal_H_long",
        "trial_windows",
        "audit_selection_summary",
    }

    trialwise_bundle = load_single_parquet_bundle(final_trialwise_alias)
    concatenated_bundle = load_single_parquet_bundle(final_concatenated_alias)
    assert set(trialwise_bundle["minimal_W"]["aggregation_mode"].unique().tolist()) == {"trialwise"}
    assert set(concatenated_bundle["minimal_W"]["aggregation_mode"].unique().tolist()) == {"concatenated"}
    assert set(trialwise_bundle["labels"]["group_id"].unique().tolist()) == {"pooled_step_nonstep"}
    assert set(concatenated_bundle["labels"]["group_id"].unique().tolist()) == {"pooled_step_nonstep"}
    assert "h_se" in trialwise_bundle["pooled_strategy_h_means"].columns
    assert "h_n" in trialwise_bundle["pooled_strategy_h_means"].columns
    assert "h_std" not in trialwise_bundle["pooled_strategy_h_means"].columns
    assert "h_se" in concatenated_bundle["pooled_strategy_h_means"].columns
    assert "h_n" in concatenated_bundle["pooled_strategy_h_means"].columns
    assert "h_std" not in concatenated_bundle["pooled_strategy_h_means"].columns

    for mode in ("trialwise", "concatenated"):
        mode_dir = run_dir / mode
        assert mode_dir.exists()
        assert (mode_dir / "clustering_audit.xlsx").exists()
        assert (mode_dir / "results_interpretation.xlsx").exists()
        assert (mode_dir / "figures" / "pooled_step_nonstep_clusters.png").exists()
        assert not (mode_dir / "parquet").exists()
        assert not list(mode_dir.rglob("*.csv"))

    trialwise_book = load_workbook(run_dir / "trialwise" / "results_interpretation.xlsx")
    try:
        assert {
            "summary",
            "clustering_meta",
            "trial_windows",
            "cluster_labels",
            "representative_W",
            "representative_H",
            "minimal_W",
            "minimal_H",
            "pooled_strategy",
            "pooled_strategy_W",
            "pooled_strategy_H",
            "table_guide",
        }.issubset(set(trialwise_book.sheetnames))
        pooled_h_sheet = trialwise_book["pooled_strategy_H"]
        pooled_h_table = pooled_h_sheet.tables["tbl_pooled_strategy_h_means"]
        min_col, min_row, max_col, _ = range_boundaries(pooled_h_table.ref)
        headers = [
            pooled_h_sheet.cell(row=min_row, column=column_index).value
            for column_index in range(min_col, max_col + 1)
        ]
        assert "h_se" in headers
        assert "h_n" in headers
        assert "h_std" not in headers
    finally:
        trialwise_book.close()


def test_trialwise_only_run_omits_concatenated_alias_and_outputs(
    repo_root: Path,
    fixture_bundle: dict[str, object],
    tmp_path: Path,
) -> None:
    """A trialwise-only run should not create concatenated-only output trees."""
    main_path = repo_root / "main.py"
    if not main_path.exists():
        pytest.xfail("main.py is not implemented yet; end-to-end contract is staged.")

    run_dir, result = _run_fixture_mode(
        repo_root,
        fixture_bundle,
        tmp_path,
        mode="trialwise",
        run_name="fixture_run_trialwise_only",
    )
    if result.returncode != 0:
        pytest.fail(
            "Fixture run failed.\n"
            f"stdout:\n{result.stdout}\n"
            f"stderr:\n{result.stderr}"
        )

    final_parquet_alias = repo_root / "outputs" / "final.parquet"
    final_trialwise_alias = repo_root / "outputs" / "final_trialwise.parquet"

    assert final_parquet_alias.exists()
    assert final_trialwise_alias.exists()
    assert (run_dir / "trialwise" / "results_interpretation.xlsx").exists()
    assert (run_dir / "trialwise" / "clustering_audit.xlsx").exists()
    assert not (run_dir / "concatenated").exists()
    assert not (run_dir / "results_interpretation.xlsx").exists()
    assert not (run_dir / "clustering_audit.xlsx").exists()
    assert not list(run_dir.rglob("*.csv"))

    trialwise_bundle = load_single_parquet_bundle(final_trialwise_alias)
    assert trialwise_bundle["source_trial_windows"].empty
    assert set(trialwise_bundle["minimal_W"]["aggregation_mode"].unique().tolist()) == {"trialwise"}
