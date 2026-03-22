"""Contract tests for the paired refilter reclustering analysis workflow."""

from __future__ import annotations

import importlib.util
import json
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd
import pytest

from src.synergy_stats.single_parquet import write_single_parquet_bundle
from tests.helpers import repo_python


REPO_ROOT = Path(__file__).resolve().parents[2]
SCRIPT_PATH = REPO_ROOT / "analysis" / "first_zero_duplicate_k_rerun" / "analyze_paired_refilter_reclustering.py"
EXPECTED_WORKBOOK_SHEETS = {"summary", "cluster_stats", "paired_detail", "table_guide"}
EXPECTED_SUMMARY_FIELDS = {
    "paired_key_n",
    "excluded_pair_key_n",
    "analysis_unit_n_postpaired",
    "paired_subset_manifest_path",
    "excluded_nonpaired_manifest_path",
    "paired_cluster_stats_csv_path",
    "paired_cluster_detail_csv_path",
    "paired_cluster_statistics_workbook_path",
    "k_selected_first_zero_duplicate",
}
EXPECTED_STATS_COLUMNS = {
    "cluster_id",
    "paired_key_n",
    "step_present_n",
    "nonstep_present_n",
    "both_present_n",
    "step_only_n",
    "nonstep_only_n",
    "both_absent_n",
    "step_presence_rate",
    "nonstep_presence_rate",
    "presence_rate_diff_step_minus_nonstep",
    "mcnemar_p",
    "mcnemar_q_bh",
    "mcnemar_note",
    "interpretation_label",
}
ALLOWED_PRESENCE_LABELS = {"both_present", "step_only", "nonstep_only", "both_absent"}


def _load_analysis_module() -> object | None:
    if not SCRIPT_PATH.exists():
        return None
    spec = importlib.util.spec_from_file_location("paired_refilter_reclustering", SCRIPT_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load analysis module: {SCRIPT_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _require_analysis_module() -> object:
    module = _load_analysis_module()
    if module is None:
        pytest.skip(f"paired analysis script is not present yet: {SCRIPT_PATH}")
    return module


def _paired_trial_rows() -> pd.DataFrame:
    rows = [
        {
            "aggregation_mode": "concatenated",
            "group_id": "pooled_step_nonstep",
            "subject": "S01",
            "velocity": 1.0,
            "trial_num": "concat_step",
            "trial_id": "S01_v1_concat_step",
            "analysis_unit_id": "S01_v1_concat_step",
            "source_trial_nums_csv": "1|2",
            "n_components": 2,
            "status": "ok",
            "analysis_selected_group_prepaired": True,
            "analysis_selected_group": True,
            "analysis_is_step": True,
            "analysis_is_nonstep": False,
            "analysis_step_class": "step",
            "analysis_pair_key": "S01|1.0",
            "analysis_is_paired_key": True,
            "analysis_pair_status": "paired_eligible",
            "analysis_window_source": "actual_step_onset",
            "analysis_window_is_surrogate": False,
        },
        {
            "aggregation_mode": "concatenated",
            "group_id": "pooled_step_nonstep",
            "subject": "S01",
            "velocity": 1.0,
            "trial_num": "concat_nonstep",
            "trial_id": "S01_v1_concat_nonstep",
            "analysis_unit_id": "S01_v1_concat_nonstep",
            "source_trial_nums_csv": "3|4",
            "n_components": 2,
            "status": "ok",
            "analysis_selected_group_prepaired": True,
            "analysis_selected_group": True,
            "analysis_is_step": False,
            "analysis_is_nonstep": True,
            "analysis_step_class": "nonstep",
            "analysis_pair_key": "S01|1.0",
            "analysis_is_paired_key": True,
            "analysis_pair_status": "paired_eligible",
            "analysis_window_source": "subject_mean_step_onset",
            "analysis_window_is_surrogate": True,
        },
        {
            "aggregation_mode": "concatenated",
            "group_id": "pooled_step_nonstep",
            "subject": "S02",
            "velocity": 1.0,
            "trial_num": "concat_step",
            "trial_id": "S02_v1_concat_step",
            "analysis_unit_id": "S02_v1_concat_step",
            "source_trial_nums_csv": "5|6",
            "n_components": 2,
            "status": "ok",
            "analysis_selected_group_prepaired": True,
            "analysis_selected_group": True,
            "analysis_is_step": True,
            "analysis_is_nonstep": False,
            "analysis_step_class": "step",
            "analysis_pair_key": "S02|1.0",
            "analysis_is_paired_key": True,
            "analysis_pair_status": "paired_eligible",
            "analysis_window_source": "actual_step_onset",
            "analysis_window_is_surrogate": False,
        },
        {
            "aggregation_mode": "concatenated",
            "group_id": "pooled_step_nonstep",
            "subject": "S02",
            "velocity": 1.0,
            "trial_num": "concat_nonstep",
            "trial_id": "S02_v1_concat_nonstep",
            "analysis_unit_id": "S02_v1_concat_nonstep",
            "source_trial_nums_csv": "7|8",
            "n_components": 2,
            "status": "ok",
            "analysis_selected_group_prepaired": True,
            "analysis_selected_group": True,
            "analysis_is_step": False,
            "analysis_is_nonstep": True,
            "analysis_step_class": "nonstep",
            "analysis_pair_key": "S02|1.0",
            "analysis_is_paired_key": True,
            "analysis_pair_status": "paired_eligible",
            "analysis_window_source": "subject_mean_step_onset",
            "analysis_window_is_surrogate": True,
        },
        {
            "aggregation_mode": "concatenated",
            "group_id": "pooled_step_nonstep",
            "subject": "S03",
            "velocity": 1.0,
            "trial_num": "concat_nonstep",
            "trial_id": "S03_v1_concat_nonstep",
            "analysis_unit_id": "S03_v1_concat_nonstep",
            "source_trial_nums_csv": "9|10",
            "n_components": 2,
            "status": "ok",
            "analysis_selected_group_prepaired": True,
            "analysis_selected_group": False,
            "analysis_is_step": False,
            "analysis_is_nonstep": True,
            "analysis_step_class": "nonstep",
            "analysis_pair_key": "S03|1.0",
            "analysis_is_paired_key": False,
            "analysis_pair_status": "nonpaired_excluded",
            "analysis_window_source": "subject_mean_step_onset",
            "analysis_window_is_surrogate": True,
        },
    ]
    return pd.DataFrame(rows)


def _paired_source_window_rows() -> pd.DataFrame:
    rows = [
        {
            "aggregation_mode": "concatenated",
            "group_id": "pooled_step_nonstep",
            "subject": "S01",
            "velocity": 1.0,
            "trial_num": "concat_step",
            "trial_id": "S01_v1_concat_step",
            "analysis_unit_id": "S01_v1_concat_step",
            "source_trial_num": 1,
            "source_trial_order": 1,
            "source_step_class": "step",
            "analysis_window_source": "actual_step_onset",
            "analysis_window_start": 10.0,
            "analysis_window_end": 20.0,
            "analysis_window_length": 10.0,
            "analysis_window_is_surrogate": False,
        },
        {
            "aggregation_mode": "concatenated",
            "group_id": "pooled_step_nonstep",
            "subject": "S01",
            "velocity": 1.0,
            "trial_num": "concat_nonstep",
            "trial_id": "S01_v1_concat_nonstep",
            "analysis_unit_id": "S01_v1_concat_nonstep",
            "source_trial_num": 2,
            "source_trial_order": 1,
            "source_step_class": "nonstep",
            "analysis_window_source": "subject_mean_step_onset",
            "analysis_window_start": 11.0,
            "analysis_window_end": 21.0,
            "analysis_window_length": 10.0,
            "analysis_window_is_surrogate": True,
        },
        {
            "aggregation_mode": "concatenated",
            "group_id": "pooled_step_nonstep",
            "subject": "S02",
            "velocity": 1.0,
            "trial_num": "concat_step",
            "trial_id": "S02_v1_concat_step",
            "analysis_unit_id": "S02_v1_concat_step",
            "source_trial_num": 3,
            "source_trial_order": 1,
            "source_step_class": "step",
            "analysis_window_source": "actual_step_onset",
            "analysis_window_start": 12.0,
            "analysis_window_end": 22.0,
            "analysis_window_length": 10.0,
            "analysis_window_is_surrogate": False,
        },
        {
            "aggregation_mode": "concatenated",
            "group_id": "pooled_step_nonstep",
            "subject": "S02",
            "velocity": 1.0,
            "trial_num": "concat_nonstep",
            "trial_id": "S02_v1_concat_nonstep",
            "analysis_unit_id": "S02_v1_concat_nonstep",
            "source_trial_num": 4,
            "source_trial_order": 1,
            "source_step_class": "nonstep",
            "analysis_window_source": "subject_mean_step_onset",
            "analysis_window_start": 13.0,
            "analysis_window_end": 23.0,
            "analysis_window_length": 10.0,
            "analysis_window_is_surrogate": True,
        },
    ]
    return pd.DataFrame(rows)


def _paired_feature_rows() -> tuple[pd.DataFrame, pd.DataFrame]:
    w_rows: list[dict[str, object]] = []
    h_rows: list[dict[str, object]] = []
    unit_specs = [
        ("S01", 1.0, "concat_step", "S01_v1_concat_step", [[1.0, 0.0], [0.0, 1.0]], [[0.9, 0.3], [0.2, 0.8]]),
        ("S01", 1.0, "concat_nonstep", "S01_v1_concat_nonstep", [[1.0, 0.0], [0.0, 1.0]], [[0.8, 0.4], [0.1, 0.9]]),
        ("S02", 1.0, "concat_step", "S02_v1_concat_step", [[0.0, 1.0], [1.0, 0.0]], [[0.1, 0.7], [0.8, 0.2]]),
        ("S02", 1.0, "concat_nonstep", "S02_v1_concat_nonstep", [[0.0, 1.0], [1.0, 0.0]], [[0.2, 0.6], [0.7, 0.3]]),
    ]
    muscles = ["M1", "M2"]
    for subject, velocity, trial_num, analysis_unit_id, component_columns, h_values in unit_specs:
        for component_index, weights in enumerate(component_columns):
            for muscle, value in zip(muscles, weights, strict=True):
                w_rows.append(
                    {
                        "aggregation_mode": "concatenated",
                        "group_id": "pooled_step_nonstep",
                        "subject": subject,
                        "velocity": velocity,
                        "trial_num": trial_num,
                        "trial_id": analysis_unit_id,
                        "analysis_unit_id": analysis_unit_id,
                        "component_index": component_index,
                        "n_components": 2,
                        "status": "ok",
                        "muscle": muscle,
                        "W_value": value,
                    }
                )
            for frame_idx, value in enumerate(h_values[component_index]):
                h_rows.append(
                    {
                        "aggregation_mode": "concatenated",
                        "group_id": "pooled_step_nonstep",
                        "subject": subject,
                        "velocity": velocity,
                        "trial_num": trial_num,
                        "trial_id": analysis_unit_id,
                        "analysis_unit_id": analysis_unit_id,
                        "component_index": component_index,
                        "n_components": 2,
                        "status": "ok",
                        "frame_idx": frame_idx,
                        "h_value": value,
                    }
                )
    return pd.DataFrame(w_rows), pd.DataFrame(h_rows)


def _paired_metadata_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "aggregation_mode": "concatenated",
                "group_id": "pooled_step_nonstep",
                "n_trials": 4,
                "n_components": 8,
                "n_clusters": 2,
                "status": "success",
                "selection_method": "first_zero_duplicate",
                "selection_status": "success_first_zero_duplicate",
                "k_gap_raw": 2,
                "k_selected": 2,
                "k_min_unique": 2,
                "duplicate_resolution": "none",
                "require_zero_duplicate_solution": True,
                "k_lb": 2,
                "repeats": 1000,
                "gap_ref_n": 500,
                "gap_ref_restarts": 100,
                "algorithm_used": "sklearn_kmeans",
                "torch_device": "",
                "torch_dtype": "",
                "random_state": 7,
                "max_iter": 50,
                "uniqueness_candidate_restarts": 25,
                "gap_by_k_json": json.dumps({"2": 1.25}, ensure_ascii=False),
                "gap_sd_by_k_json": json.dumps({"2": 0.01}, ensure_ascii=False),
                "observed_objective_by_k_json": json.dumps({"2": 42.0}, ensure_ascii=False),
                "duplicate_trial_count_by_k_json": json.dumps({"2": 0}, ensure_ascii=False),
            }
        ]
    )


def _paired_final_summary_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "aggregation_mode": "concatenated",
                "group_id": "pooled_step_nonstep",
                "n_trials": 4,
                "n_components": 8,
                "n_clusters": 2,
                "status": "success",
                "selection_method": "first_zero_duplicate",
                "selection_status": "success_first_zero_duplicate",
                "k_gap_raw": 2,
                "k_selected": 2,
                "k_min_unique": 2,
            }
        ]
    )


def _write_source_bundle(path: Path) -> Path:
    minimal_w, minimal_h = _paired_feature_rows()
    bundle = {
        "trial_windows": _paired_trial_rows(),
        "source_trial_windows": _paired_source_window_rows(),
        "minimal_W": minimal_w,
        "minimal_H_long": minimal_h,
        "metadata": _paired_metadata_frame(),
        "final_summary": _paired_final_summary_frame(),
    }
    write_single_parquet_bundle(bundle, path)
    return path


def _run_cli(repo_root: Path, source_path: Path, out_dir: Path) -> str:
    if not SCRIPT_PATH.exists():
        pytest.skip(f"paired analysis script is not present yet: {SCRIPT_PATH}")
    result = repo_python(
        repo_root,
        str(SCRIPT_PATH.relative_to(repo_root)),
        "--source-parquet",
        str(source_path),
        "--out-dir",
        str(out_dir),
        "--overwrite",
    )
    assert result.returncode == 0, result.stdout + result.stderr
    return result.stdout


def _read_json(path: Path) -> dict[str, object]:
    with path.open("r", encoding="utf-8-sig") as handle:
        return json.load(handle)


def _resolve_output_path(out_dir: Path, raw_value: object) -> Path:
    path = Path(str(raw_value))
    return path if path.is_absolute() else out_dir / path


def _validate_workbook(path: Path, module: object | None = None) -> dict[str, object]:
    validator = getattr(module, "validate_paired_cluster_statistics_workbook", None) if module is not None else None
    if callable(validator):
        return validator(path)

    workbook = load_workbook(path)
    try:
        assert EXPECTED_WORKBOOK_SHEETS.issubset(set(workbook.sheetnames))
        actual_tables: set[tuple[str, str]] = set()
        for sheet_name in EXPECTED_WORKBOOK_SHEETS:
            sheet = workbook[sheet_name]
            actual_tables.update((sheet_name, table_name) for table_name in sheet.tables.keys())
            guide_values = [sheet.cell(row=row_idx, column=1).value for row_idx in range(1, min(sheet.max_row, 24) + 1)]
            for label in ("[목적]", "[핵심 컬럼]", "[예시]"):
                assert any(isinstance(value, str) and value.startswith(label) for value in guide_values), (
                    sheet_name,
                    label,
                )

        guide_sheet = workbook["table_guide"]
        guide_table = guide_sheet.tables.get("tbl_table_guide")
        assert guide_table is not None
        documented_tables: set[tuple[str, str]] = set()
        for row in guide_sheet.iter_rows(
            min_row=2,
            max_row=guide_sheet.max_row,
            min_col=1,
            max_col=6,
            values_only=True,
        ):
            if row[0] in (None, ""):
                continue
            documented_tables.add((str(row[1]), str(row[0])))
            assert row[3] not in (None, "")
        for table_entry in sorted(actual_tables - {("table_guide", "tbl_table_guide")}):
            assert table_entry in documented_tables
    finally:
        workbook.close()
    return {"engine": "openpyxl", "workbook_path": str(path)}


def test_paired_refilter_cli_writes_expected_manifests_and_summary(repo_root: Path, tmp_path: Path) -> None:
    """The CLI should emit paired manifests and a summary with the expected bookkeeping fields."""
    _require_analysis_module()

    source_path = _write_source_bundle(tmp_path / "paired_source.parquet")
    out_dir = tmp_path / "paired_artifacts"
    _run_cli(repo_root, source_path, out_dir)

    summary = _read_json(out_dir / "summary.json")
    assert EXPECTED_SUMMARY_FIELDS.issubset(summary.keys())
    assert summary["paired_key_n"] == 2
    assert summary["excluded_pair_key_n"] == 1
    assert summary["analysis_unit_n_postpaired"] == 4
    assert summary["k_selected_first_zero_duplicate"] == 2

    paired_manifest = pd.read_csv(_resolve_output_path(out_dir, summary["paired_subset_manifest_path"]))
    excluded_manifest = pd.read_csv(_resolve_output_path(out_dir, summary["excluded_nonpaired_manifest_path"]))
    assert len(paired_manifest) == 2
    assert len(excluded_manifest) == 1
    assert "paired_key" in paired_manifest.columns
    assert "paired_key" in excluded_manifest.columns
    assert paired_manifest["paired_key"].nunique() == 2

    for field in (
        "paired_subset_manifest_path",
        "excluded_nonpaired_manifest_path",
        "paired_cluster_stats_csv_path",
        "paired_cluster_detail_csv_path",
        "paired_cluster_statistics_workbook_path",
    ):
        assert _resolve_output_path(out_dir, summary[field]).exists()

    assert (out_dir / "final.parquet").exists()
    assert (out_dir / "final_concatenated.parquet").exists()
    checksums_text = (out_dir / "checksums.md5").read_text(encoding="utf-8-sig")
    assert "analysis_methods_manifest.json" in checksums_text
    assert "paired_cluster_statistics.xlsx" in checksums_text


def test_paired_refilter_stats_use_exact_mcnemar_zero_discordant_fallback(repo_root: Path, tmp_path: Path) -> None:
    """Zero-discordant clusters should record McNemar p=1.0 with the locked note."""
    _require_analysis_module()

    source_path = _write_source_bundle(tmp_path / "paired_source.parquet")
    out_dir = tmp_path / "paired_artifacts"
    _run_cli(repo_root, source_path, out_dir)

    stats = pd.read_csv(out_dir / "paired_cluster_stats.csv")
    detail = pd.read_csv(out_dir / "paired_cluster_detail.csv")
    assert EXPECTED_STATS_COLUMNS.issubset(stats.columns)
    assert set(detail["presence_label"].dropna().unique()).issubset(ALLOWED_PRESENCE_LABELS)

    zero_discordant = stats.loc[(stats["step_only_n"] + stats["nonstep_only_n"]) == 0]
    assert not zero_discordant.empty
    assert (zero_discordant["mcnemar_p"] == 1.0).all()
    assert (zero_discordant["mcnemar_note"] == "no_discordant_pairs").all()
    assert (zero_discordant["interpretation_label"].isin({"shared_candidate", "uncertain_not_significant", "strategy_biased"})).all()

    summary = _read_json(out_dir / "summary.json")
    assert summary["paired_cluster_stats_csv_path"]
    assert summary["paired_cluster_detail_csv_path"]


def test_paired_presence_tables_keep_both_absent_rows_and_match_stats() -> None:
    """Presence detail should retain both-absent rows and reconcile exactly with cluster stats."""
    module = _require_analysis_module()

    paired_manifest = pd.DataFrame(
        [
            {
                "subject": "S01",
                "velocity": 1.0,
                "paired_key": "S01|1.0",
                "step_analysis_unit_id": "S01_step",
                "nonstep_analysis_unit_id": "S01_nonstep",
            },
            {
                "subject": "S02",
                "velocity": 1.0,
                "paired_key": "S02|1.0",
                "step_analysis_unit_id": "S02_step",
                "nonstep_analysis_unit_id": "S02_nonstep",
            },
        ]
    )
    labels_frame = pd.DataFrame(
        [
            {"analysis_unit_id": "S01_step", "cluster_id": 1},
            {"analysis_unit_id": "S01_step", "cluster_id": 2},
            {"analysis_unit_id": "S01_nonstep", "cluster_id": 2},
            {"analysis_unit_id": "S02_nonstep", "cluster_id": 2},
        ]
    )

    stats, detail = module._build_presence_tables(paired_manifest, labels_frame)
    assert "both_absent" in set(detail["presence_label"])

    for stat_row in stats.itertuples(index=False):
        cluster_detail = detail.loc[detail["cluster_id"] == stat_row.cluster_id]
        assert len(cluster_detail) == stat_row.paired_key_n
        assert int((cluster_detail["presence_label"] == "both_present").sum()) == stat_row.both_present_n
        assert int((cluster_detail["presence_label"] == "step_only").sum()) == stat_row.step_only_n
        assert int((cluster_detail["presence_label"] == "nonstep_only").sum()) == stat_row.nonstep_only_n
        assert int((cluster_detail["presence_label"] == "both_absent").sum()) == stat_row.both_absent_n


def test_exact_mcnemar_and_bh_cover_nonzero_discordant_case() -> None:
    """Non-zero discordant counts should follow the exact binomial rule and BH adjustment."""
    module = _require_analysis_module()

    p_value, note = module._exact_mcnemar_pvalue(3, 0)
    q_values = module._bh_adjust([0.25, 0.5, 1.0])

    assert p_value == pytest.approx(0.25)
    assert note == ""
    assert q_values == pytest.approx([0.75, 0.75, 1.0])


def test_paired_refilter_workbook_has_table_guide_and_validates(repo_root: Path, tmp_path: Path) -> None:
    """The workbook should reopen cleanly and expose the reading guide tables."""
    module = _require_analysis_module()

    source_path = _write_source_bundle(tmp_path / "paired_source.parquet")
    out_dir = tmp_path / "paired_artifacts"
    _run_cli(repo_root, source_path, out_dir)

    workbook_path = out_dir / "paired_cluster_statistics.xlsx"
    validation = _validate_workbook(workbook_path, module)
    assert validation["engine"] == "openpyxl"

    workbook = load_workbook(workbook_path)
    try:
        assert EXPECTED_WORKBOOK_SHEETS.issubset(set(workbook.sheetnames))
        assert "tbl_table_guide" in workbook["table_guide"].tables
    finally:
        workbook.close()
